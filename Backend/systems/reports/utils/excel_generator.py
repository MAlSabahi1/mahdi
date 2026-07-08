import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_report_excel(title, columns, rows, totals_row=None):
    """
    Generates an Excel file (in-memory) from the provided data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limits sheet name to 31 chars
    
    # Enable RTL
    ws.sheet_view.rightToLeft = True
    
    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color='000000')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark blue
    cell_font = Font(name='Arial', size=11)
    totals_font = Font(name='Arial', size=12, bold=True)
    totals_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Light gray
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. Title Row
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(columns))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = title_font
    cell.alignment = center_align

    # 2. Header Row
    current_row = 4
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
        # Adjust column width
        ws.column_dimensions[cell.column_letter].width = max(len(str(col_name)) + 5, 12)

    current_row += 1

    # 3. Data Rows
    for row_data in rows:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

    # 4. Totals Row
    if totals_row:
        for col_idx, value in enumerate(totals_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = totals_font
            cell.fill = totals_fill
            cell.alignment = center_align
            cell.border = thin_border

    # Return as bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
