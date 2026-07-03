"""
سكريبت توليد بيانات تجريبية واقعية — 500 فرد
ينشئ ملف Excel بنفس أعمدة البيانات الحقيقية تماماً
"""
import random
import openpyxl
from openpyxl.styles import Font, Alignment
import os

# ===== البيانات المرجعية =====

FIRST_NAMES = [
    'محمد', 'أحمد', 'علي', 'عبدالله', 'خالد', 'إبراهيم', 'يوسف', 'عمر',
    'حسن', 'حسين', 'صالح', 'عبدالرحمن', 'ناصر', 'فهد', 'سعيد', 'مراد',
    'طارق', 'ماجد', 'عادل', 'فيصل', 'سلطان', 'بلال', 'أنس', 'زياد',
    'منصور', 'نبيل', 'رشاد', 'جمال', 'كمال', 'هشام', 'وليد', 'رامي',
]

SECOND_NAMES = [
    'محمد', 'أحمد', 'علي', 'عبدالله', 'صالح', 'حسن', 'ناصر', 'سعيد',
    'عمر', 'خالد', 'إبراهيم', 'يوسف', 'عبدالرحمن', 'حسين', 'مراد',
]

THIRD_NAMES = [
    'محمد', 'أحمد', 'علي', 'عبدالله', 'صالح', 'ناصر', 'حسن', 'عمر',
]

FAMILY_NAMES = [
    'الحارثي', 'المراني', 'العبيدي', 'الجدعاني', 'البنوس', 'الشريف',
    'المحمدي', 'العليي', 'الزبيدي', 'النعماني', 'الكثيري', 'باعبود',
    'الصبري', 'المقطري', 'الجابري', 'العمري', 'الحميدي', 'السعدي',
    'القحطاني', 'الدوسري', 'المهري', 'الوائلي', 'الأشعري', 'الكندي',
    'البكري', 'الحضرمي', 'السلمي', 'الزهراني', 'المطيري', 'الشمراني',
]

RANKS_OFFICERS = ['لواء', 'عميد', 'عقيد', 'مقدم', 'رائد', 'نقيب', 'ملازم أول', 'ملازم ثاني']
RANKS_PERSONNEL = ['مساعد 1', 'مساعد 2', 'رقيب 1', 'رقيب 2', 'عريف', 'جندي', 'حارس']

DIRECTORATES = [
    'إدارة الأمن والنظام', 'إدارة الموارد', 'فرع القوات الخاصة',
    'فرع قوات الحماية', 'إدارة المنشآت وحماية الشخصيات', 'التوقيفات والملحق',
    'إدارة الإمداد والتموين', 'شرطة المديرية أ', 'معسكر الوحدة',
    'شرطة مديرية المديرية ب', 'شرطة مديرية حريب', 'إدارة المرور',
    'شرطة مديرية الجوبة', 'مكتب المدير العام', 'إدارة مكافحة المخدرات',
    'إدارة القيادة والسيطرة', 'إدارة الأدلة الجنائية', 'إدارة الأمن الوقائي',
    'إدارة البحث الجنائي', 'إدارة الاتصالات', 'شرطة مديرية المحافظة الأولى',
    'إدارة التدريب والتأهيل', 'شرطة مديرية جبل مراد', 'شرطة مديرية رحبه',
]

# أخطاء إملائية مقصودة لاختبار السكريبت
DIRECTORATES_MESSY = DIRECTORATES + [
    'إدرة الأمن والنظام',  # خطأ إملائي
    'فرع الامن المركزي',  # بدون همزة
    'إدارة المنشئات وحماية الشخصيات',  # إملاء خاطئ
    'قوة غير عاملة',  # حالة وليست إدارة
    'الشهداء',  # حالة وليست إدارة
    'المعسكر',  # اختصار
]

STATUSES_ACTIVE = [
    ('قوة عاملة فعلية', 'عاملين'),
    ('قوة عاملة فعلية', ''),
    ('قوة عاملة غير فعلية', 'قوة احتياط'),
    ('قوة عاملة غير فعلية', 'بدون عمل'),
    ('قوة عاملة غير فعلية', 'إجازة مفتوحه'),
]

STATUSES_TEMP_INACTIVE = [
    ('قوة غير عاملة مؤقتاً', 'غياب مستمر'),
    ('قوة غير عاملة مؤقتاً', 'كبيرسن'),
    ('قوة غير عاملة مؤقتاً', 'الجرحى'),
    ('قوة غير عاملة مؤقتاً', 'المعيات'),
    ('قوة غير عاملة مؤقتاً', 'منزل'),
    ('قوة غير عاملة مؤقتاً', 'مصدر'),
]

STATUSES_PERM_INACTIVE = [
    ('قوة غير عاملة نهائياً', 'الشهيد'),
    ('قوة غير عاملة نهائياً', 'مفصول'),
    ('قوة غير عاملة نهائياً', 'المتوفي'),
    ('قوة غير عاملة نهائياً', 'غياب مستمر'),
]

SECTIONS_REAL = ['قسم التجنيد والأرشيف', 'قسم الأمن والنظام', 'ك4/نقاط الجوبة', 'معسكر الوحدة', '']
SECTIONS_MESSY = SECTIONS_REAL + [
    'كبارين سن', 'الشهداء', 'المعيات', 'المفصولين', 'غياب مستمر',
    'قوة احتياطية', 'المصدرين', 'الجرحى',
]

WORK_TYPES = ['مهام أمنية', 'خدمة أمنية', 'تحركات', 'رئيس قسم', 'قوة احتياط', 'شهيد', 'فرار', 'معيات', 'مصدر', 'جريح', '']
POSITIONS = ['مدير إدارة', 'رئيس القسم', 'نائب مدير الإدارة', 'قائد السرية', '', '', '', '', '', '']
CATEGORIES = ['إداري', 'ميداني', '', '']
QUALIFICATIONS = ['ثانوي', 'جامعي', 'ماجستير', 'إعدادي', 'ابتدائي', '', '', '']

FORCE_TYPES = [
    'القوة الأساسي- أفراد', 'لجان أمنية- أفراد',
    'قوة- مستجدين جديد', 'قوة إداريا فقط', '', '',
]

MONTHLY_VALUES = ['نعم', 'لا', 'تم', 'لم يتم', 'صرف', 'عدم صرف', 'موجود', 'غير موجود', '']


def generate_name():
    return f"{random.choice(FIRST_NAMES)} {random.choice(SECOND_NAMES)} {random.choice(THIRD_NAMES)} {random.choice(FAMILY_NAMES)}"


def generate_mil_number(is_officer=False, force_type='basic'):
    if is_officer:
        return f"60{random.randint(10000, 99999)}"
    if force_type == 'committee':
        return f"50{random.randint(10000, 99999)}"
    return f"7{random.randint(100000, 999999)}"


def generate_national_id():
    if random.random() < 0.3:  # 30% بدون رقم وطني
        return ''
    return f"{random.randint(10000000000, 99999999999)}"


def generate_phone():
    if random.random() < 0.4:
        return ''
    return f"77{random.randint(1000000, 9999999)}"


def generate_row(idx, is_officer=False):
    # تحديد نوع القوة
    force_choice = random.choice(['basic', 'committee', 'newcomer'])
    mil = generate_mil_number(is_officer, force_choice)

    # الرتبة
    if is_officer:
        rank = random.choice(RANKS_OFFICERS)
    else:
        rank = random.choice(RANKS_PERSONNEL)

    # الرتبة الجديدة (30% لديهم)
    new_rank = ''
    if random.random() < 0.3:
        if is_officer:
            new_rank = random.choice(RANKS_OFFICERS)
        else:
            new_rank = random.choice(RANKS_PERSONNEL)

    # رقم الضابط
    officer_num = mil if is_officer else ''

    # الأرقام
    mil_correct = mil if random.random() < 0.8 else ''
    mil_old = ''
    if random.random() < 0.15:
        mil_old = f"7{random.randint(100000, 999999)}"

    # الحالة
    status_roll = random.random()
    if status_roll < 0.55:
        status_class, status_type = random.choice(STATUSES_ACTIVE)
    elif status_roll < 0.80:
        status_class, status_type = random.choice(STATUSES_TEMP_INACTIVE)
    else:
        status_class, status_type = random.choice(STATUSES_PERM_INACTIVE)

    # الإدارة (أحياناً فوضوية)
    if status_class.startswith('قوة غير عاملة'):
        directorate = random.choice(DIRECTORATES_MESSY)
    else:
        directorate = random.choice(DIRECTORATES)

    # القسم (أحياناً فوضوي)
    section = random.choice(SECTIONS_MESSY) if random.random() < 0.4 else random.choice(SECTIONS_REAL)

    # الاسم والتصحيح
    name = generate_name()
    name_correction = ''
    if random.random() < 0.1:
        name_correction = generate_name()

    # الرقم الوطني والتصحيح
    nid = generate_national_id()
    nid_correction = ''
    if nid and random.random() < 0.05:
        nid_correction = f"{random.randint(10000000000, 99999999999)}"

    # حالة الرقم الوطني
    nid_status = ''
    if nid:
        nid_status = 'موجود'
    else:
        nid_status = random.choice(['غير موجود', 'ناقص', ''])

    # تصنيف القوة
    if is_officer:
        force = 'القوة الأساسي- أفراد'  # خطأ مقصود — ضباط لكن كتبت أفراد
    else:
        force = random.choice(FORCE_TYPES)

    # المتغيرات الشهرية
    monthly = [random.choice(MONTHLY_VALUES) for _ in range(10)]

    return [
        idx,                    # م
        nid_status,             # حالة الرقم الوطني
        new_rank,               # الرتبة الجديدة
        officer_num,            # رقم الضابط
        mil_correct,            # الرقم العسكري الصحيح
        mil_old,                # الرقم العسكري القديم
        rank,                   # الرتبة
        mil,                    # الرقم العسكري
        nid,                    # الرقم الوطني
        name,                   # الاسم
        name_correction,        # تصحيح الاسم من واقع البطاقة
        nid_correction,         # تصحيح الرقم الوطني
        '',                     # نوع كشف الخدمات
        'المحافظة الأولى',                 # الوحدة
        directorate,            # الإدارة_السرية
        section,                # القسم_فرع السرية
        random.choice(POSITIONS),     # المنصب
        random.choice(WORK_TYPES),    # نوع العمل
        random.choice(CATEGORIES),    # الفئة
        status_class,           # الحالة
        status_type,            # نوع الحالة
        '',                     # عمود فارغ
        force,                  # تصنيف القوة
        random.choice(QUALIFICATIONS),  # المؤهل
        generate_phone(),       # رقم التليفون
        random.choice(['نفقات', 'لديه نفقات', '']),  # حالة النفقات
        *monthly,               # 10 أعمدة متغيرات شهرية
        random.choice(['', 'تعيين جديد', 'نقل']),  # التعيينات
        '',                     # تاريخ صدور القرار
        random.choice(['', 'لا ملاحظات', f'ملاحظة {idx}']),  # ملاحظات
        '',                     # تاريخ الميلاد
        '',                     # تاريخ الالتحاق
        '',                     # عمود فارغ
        '',                     # تاريخ التصدور إلينا
        random.choice(['', f'تنزيل {idx}']),  # ملاحظات التنزيل
    ]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'البيانات'
    ws.sheet_view.rightToLeft = True

    # الأعمدة
    headers = [
        'م', 'حالة الرقم الوطني', 'الرتبة الجديدة', 'رقم الضابط',
        'الرقم العسكري الصحيح', 'الرقم العسكري القديم', 'الرتبة',
        'الرقم العسكري', 'الرقم الوطني', 'الأسم',
        'تصحيح الأسم من واقع البطاقة', 'تصحيح الرقم الوطني',
        'نوع كشف الخدمات', 'الوحدة', 'الإدارة_السرية',
        'القسم_فرع السرية', 'المنصب', 'نوع العمل', 'الفئة',
        'الحالة', 'نوع الحالة', '',
        'تصنيف القوة', 'المؤهل', 'رقم التليفون', 'حالة النفقات',
        'متغيرات مارس 2026', 'متغيرات فبراير 2026',
        'متغيرات يناير 2026', 'متغير شهر ديسمبر',
        'متغير شهر نوفمبر', 'متغير شهر أكتوبر',
        'متغير شهر سبتمبر', 'متغير شهر أغسطس',
        'متغير شهر يونيو+يوليو', 'متغير الشهر/ابريل 2025',
        'التعيينات', 'تاريخ صدور القرار', 'ملاحظات',
        'تاريخ الميلاد', 'تاريخ الألتحاق', '',
        'تاريخ التصدور الينا', 'ملاحظات التنزيل من الخدمات',
    ]

    # كتابة العناوين
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')

    # توليد 500 صف
    for i in range(1, 501):
        is_officer = random.random() < 0.15  # 15% ضباط
        row = generate_row(i, is_officer)
        for col, val in enumerate(row, 1):
            ws.cell(row=i + 1, column=col, value=val)

    output = '/app/test_data_500.xlsx'
    wb.save(output)
    print(f'✅ تم إنشاء {output} — 500 سجل')


if __name__ == '__main__':
    main()
