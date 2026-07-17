"""
Personnel Views - واجهات API إدارة الأفراد
The personnel views use PermissionRequiredMixin + Perms.* for clean,
type-safe permission checks integrated with the ABAC data-scope engine.
"""
from rest_framework import status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import connection
from django.core.files.storage import FileSystemStorage
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter

from .models import PersonnelMaster, SuggestedCorrection
from .serializers import (
    PersonnelListSerializer, PersonnelDetailSerializer,
    PersonnelCreateSerializer, PersonnelUpdateSerializer,
    PersonnelHistorySerializer, SuggestedCorrectionSerializer,
    HistoricalMonthlyVariablesSerializer
)

# نظام الصلاحيات المتقدم — PermissionRequiredMixin مع Perms.* Registry
from infra.authorization.mixins.permission_mixin import PermissionRequiredMixin, ServicePermission
from infra.authorization.registry.permissions import Perms

# ABAC scope helpers — تصفية البيانات حسب نطاق المستخدم
from infra.security.permissions import (
    has_permission as check_perm,
    has_department_scope, filter_by_department_scope,
    check_permission, check_permission_for_department,
)

from core.base_views import BaseModelViewSet
from infra.security.dual_auth_service import DualAuthorizationService, DualAuthError
from infra.audit.models import AuditLog


@extend_schema_view(
    list=extend_schema(summary='قائمة الأفراد', tags=['personnel']),
    retrieve=extend_schema(summary='تفاصيل فرد', tags=['personnel']),
    create=extend_schema(summary='إضافة فرد جديد', tags=['personnel']),
    update=extend_schema(summary='تعديل بيانات فرد', tags=['personnel']),
    partial_update=extend_schema(summary='تعديل جزئي', tags=['personnel']),
    destroy=extend_schema(summary='حذف فرد (يتطلب تفويض مزدوج)', tags=['personnel']),
)
class PersonnelViewSet(PermissionRequiredMixin, BaseModelViewSet):
    """
    إدارة الأفراد الكاملة مع ABAC scope filtering

    نظام الصلاحيات:
      - PermissionRequiredMixin يفحص permission_map حسب الإجراء
      - Perms.* ثوابت تمنع أخطاء الإملاء
      - get_queryset يطبّق ABAC Data Scope تلقائياً

    - GET    /personnel/                       — قائمة (مع بحث وفلترة + ABAC)
    - GET    /personnel/{military_number}/     — تفاصيل
    - POST   /personnel/                       — إضافة (idempotent)
    - PUT    /personnel/{military_number}/     — تعديل (idempotent)
    - DELETE /personnel/{military_number}/     — حذف (Four-Eyes)
    - GET    /personnel/{military_number}/history/ — تاريخ Shadow Table
    """
    lookup_field = 'military_number'
    idempotent_actions = ['create', 'update', 'partial_update']

    # ——— خريطة الصلاحيات ———
    # PermissionRequiredMixin يقرأ هذه الخريطة تلقائياً بناءاً على الـ action الحالي
    permission_map = {
        'list':           Perms.PERSONNEL_VIEW,
        'retrieve':       Perms.PERSONNEL_VIEW,
        'create':         Perms.PERSONNEL_CREATE,
        'update':         Perms.PERSONNEL_EDIT,
        'partial_update': Perms.PERSONNEL_EDIT,
        'destroy':        Perms.PERSONNEL_DELETE,
        'history':        Perms.PERSONNEL_VIEW,
        'stats':          Perms.PERSONNEL_VIEW,
        'export_csv':     Perms.SHEETS_EXPORT,
        'monthly_variables': Perms.PERSONNEL_VIEW,
        'active_variables':  Perms.PERSONNEL_VIEW,
        'upsert_variable':   Perms.PERSONNEL_EDIT,
    }

    filterset_fields = [
        'security_admin', 'central_department', 'branch', 'district_police',
        'current_rank', 'current_status',
        'is_complete', 'is_data_clean',
    ]
    search_fields = ['full_name', 'military_number', 'old_military_number', 'national_id']
    ordering_fields = [
        'military_number', 'full_name', 'join_date',
        'data_quality_score', 'updated_at',
    ]
    ordering = ['military_number']

    
    def get_queryset(self):
        qs = PersonnelMaster.objects.select_related(
            'current_rank', 'central_department', 'current_status',
            'qualification',
        ).all()

        # ABAC: تصفية حسب نطاق المستخدم (استخدام security_admin)
        from infra.authorization.services.permission_service import PermissionService
        qs = PermissionService.get_scoped_queryset(
            self.request.user, qs, 'personnel.view.*'
        )
        
        # فلتر حالة الرقم الوطني (missing/valid/invalid_format/invalid_length)
        nid_status = self.request.query_params.get('national_id_status')
        if nid_status == 'missing':
            from django.db.models import Q
            qs = qs.filter(Q(national_id__isnull=True) | Q(national_id=''))
        elif nid_status == 'valid':
            qs = qs.filter(national_id__regex=r'^\d{11}$')
        elif nid_status == 'invalid':
            qs = qs.exclude(national_id__isnull=True).exclude(national_id='').exclude(national_id__regex=r'^\d{11}$')
        
        return qs
    
    def get_serializer_class(self):
        if self.action == 'list':
            return PersonnelListSerializer
        elif self.action in ('create',):
            return PersonnelCreateSerializer
        elif self.action in ('update', 'partial_update'):
            return PersonnelUpdateSerializer
        return PersonnelDetailSerializer
    
    def perform_create(self, serializer):
        personnel = serializer.save()
        AuditLog.objects.create(
            user=self.request.user,
            action='CREATE',
            model_name='PersonnelMaster',
            object_id=personnel.military_number,
            new_data=PersonnelListSerializer(personnel).data,
            ip_address=self._get_ip(),
        )
    
    def perform_update(self, serializer):
        old_data = PersonnelListSerializer(self.get_object()).data
        personnel = serializer.save()
        AuditLog.objects.create(
            user=self.request.user,
            action='UPDATE',
            model_name='PersonnelMaster',
            object_id=personnel.military_number,
            old_data=old_data,
            new_data=PersonnelListSerializer(personnel).data,
            ip_address=self._get_ip(),
        )
    
    def destroy(self, request, *args, **kwargs):
        """حذف يتطلب تفويض مزدوج"""
        personnel = self.get_object()
        
        try:
            service = DualAuthorizationService(request.user)
            reason = request.data.get('reason', '')
            if not reason:
                return Response(
                    {'success': False, 'error': 'يجب تحديد سبب الحذف'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            req = service.create_request(
                action_type='DELETE_PERSONNEL',
                target_object_type='PersonnelMaster',
                target_object_id=personnel.military_number,
                reason=reason,
            )
            
            return Response({
                'success': True,
                'requires_dual_auth': True,
                'dual_auth_request_id': str(req.pk),
                'message': 'تم إنشاء طلب تفويض مزدوج للحذف. '
                          'بانتظار موافقة مدير آخر.',
            })
        except DualAuthError as e:
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @extend_schema(summary='حقول التصدير المتاحة وتكويناتها', tags=['personnel'])
    @action(detail=False, methods=['get'], url_path='export-fields')
    def export_fields(self, request, *args, **kwargs):
        from django.apps import apps
        PersonnelMaster = apps.get_model('personnel', 'PersonnelMaster')
        
        def get_label(field_name):
            try:
                return str(PersonnelMaster._meta.get_field(field_name).verbose_name)
            except Exception:
                if field_name == 'qualification':
                    return 'المؤهل الدراسي'
                elif field_name == 'current_rank':
                    return 'الرتبة الحالية'
                elif field_name == 'current_status':
                    return 'الحالة الخدمية الحالية'
                elif field_name == 'force_classification':
                    return 'تصنيف القوة'
                elif field_name == 'job_title':
                    return 'نوع العمل'
                return field_name
        
        data = {
            'identity': [
                {'field': 'current_rank', 'label': get_label('current_rank'), 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': True},
                {'field': 'military_number', 'label': get_label('military_number'), 'locked': True, 'alwaysLocked': True, 'exportable': True, 'alwaysExportable': True},
                {'field': 'national_id', 'label': get_label('national_id'), 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'full_name', 'label': get_label('full_name'), 'locked': True, 'alwaysLocked': True, 'exportable': True, 'alwaysExportable': True},
            ],
            'structure': [
                {'field': 'security_admin', 'label': 'الوحدة', 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'central_department_or_branch', 'label': 'الإدارة_السرية', 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'district_police_or_division', 'label': 'القسم_فرع السرية', 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'position', 'label': get_label('position'), 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'job_title', 'label': get_label('job_title'), 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'category', 'label': 'الفئة', 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
            ],
            'statusAndDecisions': [
                {'field': 'current_status', 'label': 'نوع الحالة', 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': True},
                {'field': 'pseudo_status_type', 'label': 'الحالة', 'locked': False, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': True},
                {'field': 'force_classification', 'label': get_label('force_classification'), 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'qualification', 'label': get_label('qualification'), 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'phone_number', 'label': get_label('phone_number'), 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'expense_status', 'label': get_label('expense_status'), 'locked': True, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': False},
                {'field': 'pseudo_monthly_var', 'label': 'متغيرات الشهر', 'locked': False, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': True},
                {'field': 'pseudo_notes', 'label': 'ملاحظات', 'locked': False, 'alwaysLocked': False, 'exportable': True, 'alwaysExportable': True},
            ]
        }
        return Response({'success': True, 'groups': data})

    @extend_schema(summary='تصدير قائمة الأفراد', tags=['personnel'])
    @action(detail=False, methods=['get'])
    def export_csv(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Protection, Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from django.apps import apps
        import io

        columns_param = request.query_params.get('columns', '')
        locked_param = request.query_params.get('locked_columns', '')
        
        security_admins_param = request.query_params.get('security_admins', '')
        central_deps_param = request.query_params.get('central_departments', '')
        branches_param = request.query_params.get('branches', '')
        district_polices_param = request.query_params.get('district_polices', '')
        statuses_param = request.query_params.get('statuses', '')
        
        split_by = request.query_params.get('split_by', '')
        
        locked_fields = [f.strip() for f in locked_param.split(',') if f.strip()]
        
        qs = self.get_queryset()
        
        if security_admins_param:
            ids = [int(x) for x in security_admins_param.split(',') if x.strip().isdigit()]
            if ids:
                qs = qs.filter(security_admin_id__in=ids)
        if central_deps_param:
            ids = [int(x) for x in central_deps_param.split(',') if x.strip().isdigit()]
            if ids:
                qs = qs.filter(central_department_id__in=ids)
        if branches_param:
            ids = [int(x) for x in branches_param.split(',') if x.strip().isdigit()]
            if ids:
                qs = qs.filter(branch_id__in=ids)
        if district_polices_param:
            ids = [int(x) for x in district_polices_param.split(',') if x.strip().isdigit()]
            if ids:
                qs = qs.filter(district_police_id__in=ids)
        if statuses_param:
            ids = [int(x) for x in statuses_param.split(',') if x.strip().isdigit()]
            if ids:
                qs = qs.filter(current_status_id__in=ids)
                
        qs = self.filter_queryset(qs)
        
        PersonnelMaster = apps.get_model('personnel', 'PersonnelMaster')
        
        def get_field_label(field_name):
            try:
                return str(PersonnelMaster._meta.get_field(field_name).verbose_name)
            except Exception:
                if field_name == 'qualification':
                    return 'المؤهل الدراسي'
                elif field_name == 'current_rank':
                    return 'الرتبة الحالية'
                elif field_name == 'current_status':
                    return 'الحالة الخدمية الحالية'
                elif field_name == 'force_classification':
                    return 'تصنيف القوة'
                elif field_name == 'job_title':
                    return 'نوع العمل'
                return field_name

        supported_fields = [
            'military_number', 'old_military_number', 'full_name', 'national_id',
            'phone_number', 'birth_date', 'qualification', 'security_admin',
            'central_department', 'branch', 'district_police', 'division',
            'unit', 'force_classification', 'job_title', 'position',
            'current_rank', 'current_status', 'join_date', 'decision_date',
            'transfer_date', 'expense_status', 'notes'
        ]
        
        if columns_param:
            export_fields = [f.strip() for f in columns_param.split(',') if f.strip() in supported_fields]
        else:
            export_fields = supported_fields
            
        if 'military_number' not in export_fields:
            export_fields.insert(0, 'military_number')
        if 'full_name' not in export_fields:
            idx = export_fields.index('military_number') + 1
            export_fields.insert(idx, 'full_name')

        grouped_data = {}
        if split_by == 'official_monthly_roster':
            grouped_data = {
                'قوة عاملة': [p for p in qs if p.current_status and p.current_status.classification in ('active_full', 'active_part')],
                'غير عاملة': [p for p in qs if p.current_status and p.current_status.classification in ('inactive_temp', 'inactive_perm')],
                'كاملة': list(qs),
                'غياب': [p for p in qs if p.current_status and 'غياب' in p.current_status.name]
            }
        elif split_by in ['security_admin', 'central_department', 'branch']:
            for person in qs:
                group_obj = getattr(person, split_by, None)
                group_name = group_obj.name if group_obj else 'غير محدد'
                sheet_name = group_name[:30].replace(':', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace('/', '').replace('\\', '')
                if not sheet_name:
                    sheet_name = 'غير محدد'
                if sheet_name not in grouped_data:
                    grouped_data[sheet_name] = []
                grouped_data[sheet_name].append(person)
        else:
            grouped_data['الأفراد'] = list(qs)

        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        fill_header_locked = PatternFill(start_color='B22222', end_color='B22222', fill_type='solid')
        fill_header_editable = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
        
        font_cell = Font(name='Calibri', size=11)
        fill_cell_locked = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        
        border_thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        def get_val(person, field_name):
            val = getattr(person, field_name, None)
            if val is None:
                return ''
            if hasattr(val, 'name'):
                return val.name
            from datetime import date
            if isinstance(val, date):
                return val.strftime('%Y-%m-%d')
            return str(val)

        for sheet_name, people in grouped_data.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.views.sheetView[0].rightToLeft = True

            for col_idx, field in enumerate(export_fields, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = get_field_label(field)
                cell.font = font_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
                
                is_locked = field in locked_fields
                cell.fill = fill_header_locked if is_locked else fill_header_editable

            for row_idx, person in enumerate(people, start=2):
                for col_idx, field in enumerate(export_fields, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = get_val(person, field)
                    cell.font = font_cell
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border_thin
                    
                    is_locked = field in locked_fields
                    if is_locked:
                        cell.protection = Protection(locked=True)
                        cell.fill = fill_cell_locked
                    else:
                        cell.protection = Protection(locked=False)

            ws.protection.sheet = True
            ws.protection.enable()

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        if not wb.sheetnames:
            ws = wb.create_sheet(title="لا توجد بيانات")
            ws.views.sheetView[0].rightToLeft = True

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="personnel.xlsx"'
        return response
    
    @extend_schema(
        summary='تاريخ الفرد من جدول الظل',
        tags=['personnel'],
        responses={200: PersonnelHistorySerializer(many=True)},
    )
    @action(detail=True, methods=['get'])
    def history(self, request, military_number=None):
        """عرض التاريخ الكامل من Shadow Table"""
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT history_id, history_action, history_user,
                       history_timestamp, history_version,
                       military_number, full_name, national_id,
                       central_department_id, current_rank_id,
                       current_status_id
                FROM personnel_master_history
                WHERE military_number = %s
                ORDER BY history_timestamp DESC
                LIMIT 100
                """,
                [military_number]
            )
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        serializer = PersonnelHistorySerializer(rows, many=True)
        return Response({
            'success': True,
            'count': len(rows),
            'data': serializer.data,
        })
    
    @extend_schema(
        summary='المتغيرات الشهرية للفرد',
        tags=['personnel'],
        responses={200: HistoricalMonthlyVariablesSerializer(many=True)},
    )
    @action(detail=True, methods=['get'], url_path='monthly-variables')
    def monthly_variables(self, request, military_number=None):
        """عرض المتغيرات الشهرية للفرد"""
        from .models import HistoricalMonthlyVariables
        from .serializers import HistoricalMonthlyVariablesSerializer
        personnel = self.get_object()
        qs = HistoricalMonthlyVariables.objects.filter(personnel=personnel).order_by('-month')
        serializer = HistoricalMonthlyVariablesSerializer(qs, many=True)
        return Response({
            'success': True,
            'count': qs.count(),
            'data': serializer.data,
        })

    @extend_schema(
        summary='جلب الأفراد مع متغيراتهم لشهر محدد (لشبكة البيانات)',
        parameters=[OpenApiParameter(name='month', description='مثال: 2026-05', type=str)],
        tags=['personnel']
    )
    @action(detail=False, methods=['get'], url_path='active-variables')
    def active_variables(self, request):
        month = request.query_params.get('month')
        if not month:
            return Response({"error": "يجب تحديد الشهر parameter: month"}, status=400)

        from .models import HistoricalMonthlyVariables
        from django.db.models import Prefetch
        from .serializers import PersonnelActiveVariableSerializer
        
        prefetch = Prefetch(
            'monthly_variables',
            queryset=HistoricalMonthlyVariables.objects.filter(month=month),
            to_attr='prefetched_active_variables'
        )
        
        qs = self.get_queryset().select_related('current_rank', 'central_department').prefetch_related(prefetch)
        page = self.paginate_queryset(qs)
        serializer = PersonnelActiveVariableSerializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @extend_schema(summary='إضافة أو تعديل متغير شهري من شبكة البيانات (Inline Edit)', tags=['personnel'])
    @action(detail=False, methods=['post', 'patch'], url_path='upsert-variable')
    def upsert_variable(self, request):
        from .serializers import ActiveVariableUpsertSerializer
        from .models import HistoricalMonthlyVariables
        from django.shortcuts import get_object_or_404
        
        serializer = ActiveVariableUpsertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        personnel = get_object_or_404(self.get_queryset(), military_number=data['military_number'])
        
        if not data['value'].strip():
            HistoricalMonthlyVariables.objects.filter(personnel=personnel, month=data['month']).delete()
        else:
            HistoricalMonthlyVariables.objects.update_or_create(
                personnel=personnel,
                month=data['month'],
                defaults={'variable_value': data['value'], 'source': 'manual'}
            )
        
        return Response({"success": True})

    @extend_schema(
        summary='إحصائيات الأفراد',
        tags=['personnel'],
    )
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """إحصائيات سريعة"""
        qs = self.get_queryset()
        from django.db.models import Count, Avg
        
        stats = {
            'total': qs.count(),
            'complete': qs.filter(is_complete=True).count(),
            'incomplete': qs.filter(is_complete=False).count(),
            'clean_data': qs.filter(is_data_clean=True).count(),
            'avg_quality': qs.aggregate(
                avg=Avg('data_quality_score')
            )['avg'],
            'by_status': list(
                qs.values('current_status__name')
                .annotate(count=Count('military_number'))
                .order_by('-count')
            ),
            'by_department': list(
                qs.values('central_department__name')
                .annotate(count=Count('military_number'))
                .order_by('-count')
            ),
        }
        
        return Response({'success': True, 'data': stats})
    
    def _get_ip(self):
        xff = self.request.META.get('HTTP_X_FORWARDED_FOR')
        return xff.split(',')[0].strip() if xff else self.request.META.get('REMOTE_ADDR')

class LegacyImportView(APIView):
    """
    مسار مخصص لمدير النظام حصراً لرفع البيانات الخام التاريخية
    """
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
    parser_classes = [MultiPartParser, FormParser]

    @extend_schema(
        summary='رفع ملف البيانات الأصلية التأسيسية',
        description='API مخصص لرفع الملف المليء بالفوضى لأول مرة. يتطلب صلاحيات سوبر يوزر ويعمل من خلال Celery Task',
        tags=['admin', 'import']
    )
    def post(self, request):
        if 'file' not in request.FILES:
            return Response({'success': False, 'error': 'يجب إرفاق ملف إكسل'}, status=status.HTTP_400_BAD_REQUEST)
        
        file = request.FILES['file']
        is_dry_run = request.data.get('dry_run', 'true').lower() == 'true'

        from django.core.files.storage import default_storage
        import os
        from django.conf import settings
        
        # حفظ الملف في مجلد مؤقت
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'temp_imports'))
        if not fs.exists(''):
            os.makedirs(fs.location, exist_ok=True)
            
        filename = fs.save(file.name, file)
        file_path = fs.path(filename)
        
        # تشغيل Celery Task
        from .tasks import process_legacy_import_task
        task = process_legacy_import_task.delay(file_path=file_path, dry_run=is_dry_run)
        
        return Response({
            'success': True,
            'message': 'تم استلام الملف وجاري المعالجة النظيفة في الخلفية',
            'task_id': task.id,
            'dry_run': is_dry_run
        }, status=status.HTTP_202_ACCEPTED)
# ═══════════════════════════════════════════════════
# Endpoints: إدارة الرقم الوطني
# ═══════════════════════════════════════════════════

class CheckNationalIdView(APIView):
    """
    فحص فوري للرقم الوطني — يُفوّض لـ PersonnelService.check_national_id
    GET /api/v1/personnel/check-national-id/?value=01010409070&exclude=7348799
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .services import PersonnelService
        
        value = request.query_params.get('value', '').strip()
        exclude_mil = request.query_params.get('exclude', '')
        
        if not value:
            return Response({'error': 'الرجاء إرسال الرقم الوطني في الباراميتر value'},
                            status=status.HTTP_400_BAD_REQUEST)
        
        result = PersonnelService.check_national_id(value, exclude_mil or None)
        return Response(result)


class CheckMilitaryNumberView(APIView):
    """
    فحص فوري للرقم العسكري — يُفوّض لـ PersonnelService.check_military_number
    GET /api/v1/personnel/check-military-number/?value=6012345&exclude=7348799
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        from .services import PersonnelService
        
        value = request.query_params.get('value', '').strip()
        exclude_mil = request.query_params.get('exclude', '')
        
        if not value:
            return Response({'error': 'الرجاء إرسال الرقم العسكري في الباراميتر value'},
                            status=status.HTTP_400_BAD_REQUEST)
        
        result = PersonnelService.check_military_number(value, exclude_mil or None)
        return Response(result)


class UpdateNationalIdView(APIView):
    """
    تحديث الرقم الوطني — يُفوّض لـ PersonnelService.
    - مدير/رئيس خدمات: تحديث مباشر
    - مستخدم عادي: إنشاء طلب تصحيح
    
    POST /api/v1/personnel/<military_number>/update-national-id/
    """
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request, military_number):
        from .services import PersonnelService
        
        user = request.user
        try:
            personnel = PersonnelMaster.objects.get(military_number=military_number)
        except PersonnelMaster.DoesNotExist:
            return Response({'success': False, 'error': 'الفرد غير موجود'},
                            status=status.HTTP_404_NOT_FOUND)
        
        new_national_id = request.data.get('national_id', '').strip()
        document_ids = request.data.get('document_ids', [])
        
        # دعم الطريقة القديمة
        front_doc_id = request.data.get('front_document_id')
        back_doc_id = request.data.get('back_document_id')
        if front_doc_id and back_doc_id and not document_ids:
            document_ids = [front_doc_id, back_doc_id]
        
        ip_address = (
            request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
            or request.META.get('REMOTE_ADDR')
        )
        
        is_admin = user.is_superuser or (
            hasattr(user, 'role') and user.role and
            user.role.name in ('مدير النظام', 'رئيس الخدمات')
        )
        
        try:
            if is_admin:
                PersonnelService.update_national_id(
                    personnel, new_national_id, document_ids,
                    user=user, ip_address=ip_address,
                )
                return Response({
                    'success': True,
                    'message': 'تم تحديث الرقم الوطني وربط المرفقات بنجاح',
                    'mode': 'direct_save',
                    'updated_by': user.username,
                })
            else:
                correction = PersonnelService.request_national_id_correction(
                    personnel, new_national_id, document_ids,
                    user=user, ip_address=ip_address,
                )
                return Response({
                    'success': True,
                    'message': 'تم إرسال طلب تعديل الرقم الوطني إلى رئيس الخدمات',
                    'mode': 'correction_request',
                    'correction_id': correction.pk,
                    'requested_by': user.username,
                }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


# SuggestedCorrection already imported at top from .models
from django.utils import timezone

class SuggestedCorrectionViewSet(BaseModelViewSet):
    """
    ViewSet لإدارة طلبات التصحيح (SuggestedCorrection).

    ─── للمطور ──────────────────────────────────────────────────
    GET    /corrections/                    → قائمة الطلبات
    POST   /corrections/                    → إنشاء طلب فردي
    GET    /corrections/{id}/               → تفاصيل طلب
    POST   /corrections/{id}/approve/       → موافقة فردية
    POST   /corrections/{id}/reject/        → رفض فردي
    POST   /corrections/approve_batch/      → موافقة جماعية (مع مذكرة الوزارة)
    POST   /corrections/reject_batch/       → رفض جماعي
    POST   /corrections/bulk-create/        → إنشاء طلبات اسم جماعية
    GET    /corrections/needs-correction/   → أفراد يحتاجون تصحيح اسم
    GET    /corrections/export/             → تصدير نموذج 23
    ─────────────────────────────────────────────────────────────
    ABAC Scope:
      - superuser / staff → يرى كل الطلبات
      - مستخدم عادي      → يرى فقط طلبات إدارة الأمن التابعة له
    """
    queryset = SuggestedCorrection.objects.select_related(
        'personnel', 'personnel__current_rank',
        'personnel__central_department',
        'security_admin',
        'requested_by', 'reviewed_by',
        'supporting_document', 'approval_document',
    )
    serializer_class = SuggestedCorrectionSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """
        ABAC Scope: المستخدم يرى فقط طلبات نطاق إدارته.
        Superuser / staff يرون كل شيء.
        """
        qs = super().get_queryset()

        # ── تصفية النطاق الأمني ──
        user = self.request.user
        if not (user.is_superuser or user.is_staff):
            # جلب security_admin من ملف المستخدم إن وُجد
            security_admin_id = getattr(
                getattr(user, 'authz_profile', None),
                'security_admin_id', None
            )
            if security_admin_id:
                qs = qs.filter(security_admin_id=security_admin_id)
            else:
                # المستخدم ليس له إدارة أمن → لا يرى شيئاً
                qs = qs.none()

        # ── Query Params حسب API Contract ──
        params = self.request.query_params

        status_param = params.get('status')
        if status_param:
            qs = qs.filter(status=status_param)

        type_param = params.get('type')
        if type_param:
            qs = qs.filter(correction_type=type_param)

        personnel_param = params.get('personnel')
        if personnel_param:
            qs = qs.filter(personnel__military_number=personnel_param)

        return qs.order_by('-requested_at')

    def create(self, request, *args, **kwargs):
        """
        إنشاء طلب تصحيح فردي — الـ Serializer يتولى كل التحقق والحفظ.

        --- للمطور ---
        POST /api/v1/personnel/corrections/
        Body: { personnel_military_number_input, correction_type,
                field_name, old_value, new_value,
                supporting_document (UUID), notes, metadata }
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        correction = serializer.save()
        return Response(
            self.get_serializer(correction).data,
            status=status.HTTP_201_CREATED
        )



    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """قبول طلب التصحيح وتحديث البيانات الرسمية باستخدام الخدمات لضمان التدقيق"""
        correction = self.get_object()
        if correction.status != 'pending':
            return Response({'success': False, 'error': 'الطلب ليس في حالة الانتظار'}, status=status.HTTP_400_BAD_REQUEST)
        
        personnel = correction.personnel
        if personnel:
            from systems.personnel.services import PersonnelService
            from systems.services.attachment_service import AttachmentService
            try:
                # جمع مرفقات السياق المرتبطة بالطلب
                linked_docs = AttachmentService.get_by_context(
                    'SuggestedCorrection', correction.pk, status='temp'
                )
                doc_ids = list(linked_docs.values_list('id', flat=True))
                
                if correction.field_name == 'full_name':
                    # ── تصحيح الاسم عبر الخدمة المركزية ──
                    PersonnelService.correct_name(
                        personnel,
                        correction.new_value,
                        document=correction.supporting_document,
                        document_ids=doc_ids if doc_ids else None,
                        user=request.user,
                    )
                elif correction.field_name == 'national_id':
                    PersonnelService.update_personnel(
                        personnel, {'national_id': correction.new_value}, user=request.user
                    )
                    # تثبيت المرفقات
                    if doc_ids:
                        AttachmentService.commit_documents(doc_ids)
                    
                elif correction.field_name == 'military_number':
                    PersonnelService.update_personnel(
                        personnel, {'military_number': correction.new_value}, user=request.user
                    )
                    if doc_ids:
                        AttachmentService.commit_documents(doc_ids)
                elif correction.field_name in ['rank', 'current_rank_id']:
                    from core.models import Rank
                    rank_obj = Rank.objects.get(pk=int(correction.new_value))
                    PersonnelService.update_personnel(
                        personnel, {'current_rank': rank_obj}, user=request.user
                    )
                    if doc_ids:
                        AttachmentService.commit_documents(doc_ids)
            except Exception as e:
                return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
            
        correction.status = 'approved'
        correction.reviewed_by = request.user
        correction.reviewed_at = timezone.now()
        correction.save()
        
        return Response({
            'success': True,
            'message': 'تم قبول طلب التصحيح وتحديث السجلات بنجاح',
            'approved_by': request.user.username,
            'requested_by': correction.requested_by.username if correction.requested_by else None,
        })

    @action(detail=False, methods=['post'], url_path='approve_batch')
    def approve_batch(self, request):
        """
        موافقة جماعية على طلبات التصحيح.
        يتطلب: correction_ids, memo_document_id
        """
        correction_ids = request.data.get('correction_ids', [])
        memo_document_id = request.data.get('memo_document_id')

        if not correction_ids:
            return Response({'success': False, 'error': 'لم يتم تحديد طلبات للموافقة عليها'}, status=status.HTTP_400_BAD_REQUEST)
        if not memo_document_id:
            return Response({'success': False, 'error': 'يجب إرفاق المذكرة الوزارية (memo_document_id)'}, status=status.HTTP_400_BAD_REQUEST)

        from systems.personnel.services import PersonnelService
        try:
            result = PersonnelService.bulk_approve_corrections(
                correction_ids=correction_ids,
                memo_document_ids=[memo_document_id],
                user=request.user
            )
            return Response({'success': True, 'data': result})
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='reject_batch')
    def reject_batch(self, request):
        """
        رفض جماعي لطلبات التصحيح.

        --- لمطور الفرونت إند ---
        POST /api/v1/personnel/corrections/reject_batch/
        Content-Type: application/json

        الجسم (Body):
        {
          "correction_ids": [1, 2, 5],   // إلزامي — معرّفات الطلبات
          "reason": "سبب الرفض",        // إلزامي
          "clear_name": false            // اختياري — true = الوزارة أكدت صحة الاسم
        }

        الاستجابة (Response):
        { "success": true, "rejected_count": 3 }
        """
        correction_ids = request.data.get('correction_ids', [])
        reason = request.data.get('reason', '')
        clear_name = bool(request.data.get('clear_name', False))

        if not correction_ids:
            return Response(
                {'success': False, 'error': 'لم يتم تحديد طلبات للرفض'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not reason:
            return Response(
                {'success': False, 'error': 'يجب ذكر سبب الرفض'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from systems.personnel.services import PersonnelService
        try:
            count = PersonnelService.bulk_reject_corrections(
                correction_ids=correction_ids,
                reason=reason,
                user=request.user,
                clear_name=clear_name,
            )
            return Response({'success': True, 'rejected_count': count})
        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """
        رفض طلب تصحيح فردي.

        --- لمطور الفرونت إند ---
        POST /api/v1/personnel/corrections/{id}/reject/
        Content-Type: application/json

        الجسم (Body):
        {
          "reason": "سبب الرفض",        // إلزامي
          "clear_name": false            // اختياري — true يُفرّغ corrected_name
                                         //   استخدم true فقط عندما تؤكد الوزارة
                                         //   أن الاسم صحيح ولا حاجة لإعادة تقديم
        }

        الاستجابة (Response):
        { "success": true, "message": "..." }
        """
        correction = self.get_object()
        if correction.status != 'pending':
            return Response(
                {'success': False, 'error': 'الطلب ليس في حالة الانتظار'},
                status=status.HTTP_400_BAD_REQUEST
            )

        reason = request.data.get('reason', '')
        if not reason:
            return Response(
                {'success': False, 'error': 'يجب ذكر سبب الرفض'},
                status=status.HTTP_400_BAD_REQUEST
            )
        clear_name = bool(request.data.get('clear_name', False))

        from systems.personnel.services import PersonnelService
        PersonnelService.bulk_reject_corrections(
            correction_ids=[correction.pk],
            reason=reason,
            user=request.user,
            clear_name=clear_name,
        )
        return Response({'success': True, 'message': 'تم رفض طلب التصحيح'})

    # ═══════════════════════════════════════════════════
    # طلب تصحيح جماعي (Bulk Create)
    # ═══════════════════════════════════════════════════
    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        """
        إنشاء طلبات تصحيح اسم لمجموعة أفراد دفعة واحدة.
        يُفوّض لـ PersonnelService.submit_name_corrections_batch
        """
        from .services import PersonnelService
        
        corrections_data = request.data.get('corrections', [])
        document_ids = request.data.get('document_ids', [])
        if not corrections_data:
            return Response(
                {'success': False, 'error': 'لم يتم إرسال أي بيانات تصحيح'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not document_ids:
            return Response(
                {'success': False, 'error': 'يجب إرفاق مستند داعم (document_ids) لتصحيح الأسماء'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = PersonnelService.submit_name_corrections_batch(
            corrections_data, document_ids=document_ids, user=request.user,
        )
        
        return Response({
            'success': True,
            'message': f'تم إنشاء {len(result["created"])} طلب تصحيح',
            'created_count': len(result['created']),
            'error_count': len(result['errors']),
            'created': result['created'],
            'errors': result['errors'],
        }, status=status.HTTP_201_CREATED)

    # ═══════════════════════════════════════════════════
    # أفراد يحتاجون تصحيح اسم
    # ═══════════════════════════════════════════════════
    @action(detail=False, methods=['get'], url_path='needs-correction')
    def needs_correction(self, request):
        """
        قائمة الأفراد الذين لديهم طلبات تصحيح اسم معلقة (بانتظار التصحيح).
        يستخدمه رئيس الخدمات لعرض من يحتاج تصحيح.

        مصدر البيانات: SuggestedCorrection (correction_type='name_correction', status='pending')
        لا يوجد corrected_name في قاعدة البيانات بعد الآن.
        """
        qs = SuggestedCorrection.objects.filter(
            correction_type='name_correction',
            status='pending',
            personnel__isnull=False,
        ).select_related(
            'personnel', 'personnel__current_rank', 'personnel__central_department'
        ).values(
            'id',
            'personnel__military_number',
            'personnel__full_name',
            'new_value',
            'requested_at',
            'personnel__current_rank__name',
            'personnel__central_department__name',
        )

        data = [{
            'correction_id': r['id'],
            'military_number': r['personnel__military_number'],
            'current_name': r['personnel__full_name'],
            'corrected_name': r['new_value'],
            'requested_at': r['requested_at'],
            'rank': r['personnel__current_rank__name'],
            'department': r['personnel__central_department__name'],
        } for r in qs]

        return Response({
            'success': True,
            'count': len(data),
            'data': data,
        })

    # ═══════════════════════════════════════════════════
    # تصدير كشف التصحيح — نموذج (23)
    # ═══════════════════════════════════════════════════
    @action(detail=False, methods=['get'], url_path='export')
    def export_corrections(self, request):
        """
        تصدير كشف طلبات التصحيح المعلقة (نموذج 23) بصيغة JSON جاهزة للتحويل إلى Excel/PDF.
        
        Params: ?status=pending&type=name_correction
        """
        qs = self.get_queryset()
        
        # فلاتر
        corr_status = request.query_params.get('status', 'pending')
        corr_type = request.query_params.get('type')
        
        qs = qs.filter(status=corr_status)
        if corr_type:
            qs = qs.filter(correction_type=corr_type)
        
        qs = qs.select_related('personnel', 'personnel__current_rank')
        
        rows = []
        for i, c in enumerate(qs, 1):
            p = c.personnel
            rows.append({
                'seq': i,
                'rank': p.current_rank.name if p and p.current_rank else '',
                'military_number': p.military_number if p else '',
                'correct_name': c.new_value,
                'wrong_name': c.old_value,
                'correction_target': c.field_name,
                'notes': '',
                'correction_id': c.pk,
                'status': c.get_status_display(),
                'requested_by': c.requested_by.username if c.requested_by else '',
                'requested_at': c.requested_at.isoformat() if c.requested_at else '',
            })
        
        return Response({
            'success': True,
            'title': 'كشف بأسماء المطلوب تصحيح أسماؤهم — نموذج رقم (23)',
            'count': len(rows),
            'columns': [
                {'key': 'seq', 'label': 'م'},
                {'key': 'rank', 'label': 'الرتبة'},
                {'key': 'military_number', 'label': 'الرقم العسكري'},
                {'key': 'correct_name', 'label': 'الاسم الصحيح'},
                {'key': 'wrong_name', 'label': 'الاسم الخطأ'},
                {'key': 'correction_target', 'label': 'المطلوب تصحيحه'},
                {'key': 'notes', 'label': 'ملاحظات'},
            ],
            'data': rows,
        })


class RankSettlementView(APIView):
    """
    تسوية الرتب لعدة أفراد دفعة واحدة
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """جلب الأفراد الذين لديهم رتبة جديدة معلقة بانتظار التسوية"""
        from .models import PersonnelMaster
        qs = PersonnelMaster.objects.select_related(
            'current_rank', 'pending_rank', 'central_department'
        ).filter(pending_rank__isnull=False)
        data = [{
            'military_number': p.military_number,
            'full_name': p.full_name,
            'current_rank_name': p.current_rank.name if p.current_rank else 'غير محدد',
            'pending_rank_name': p.pending_rank.name if p.pending_rank else 'غير محدد',
            'department_name': p.central_department.name if p.central_department else 'غير محدد',
        } for p in qs]
        return Response({'success': True, 'data': data})

    def post(self, request):
        """تسوية الرتب للمجموعة المحددة مع تسجيل حركة التعديل في سجلات التدقيق"""
        military_numbers = request.data.get('military_numbers', [])
        if not military_numbers:
            return Response({'success': False, 'error': 'لم يتم تحديد أي أفراد للتسوية'}, status=status.HTTP_400_BAD_REQUEST)
        
        from django.db import transaction
        from core.models import AuditLog
        
        with transaction.atomic():
            qs = PersonnelMaster.objects.select_related('pending_rank').filter(
                military_number__in=military_numbers, 
                pending_rank__isnull=False
            )
            
            audit_logs = []
            updated_personnel = []
            count = 0
            
            for p in qs:
                old_data = {
                    'full_name': p.full_name,
                    'national_id': p.national_id,
                    'phone_number': p.phone_number,
                }
                
                new_data = {
                    'current_rank': p.pending_rank.pk if p.pending_rank else None,
                    'pending_rank': None,
                }
                
                p.current_rank = p.pending_rank
                p.pending_rank = None
                updated_personnel.append(p)
                
                audit_logs.append(AuditLog(
                    user=request.user,
                    action='UPDATE',
                    model_name='PersonnelMaster',
                    object_id=p.military_number,
                    old_data=old_data,
                    new_data=new_data,
                ))
                count += 1
                
            if updated_personnel:
                PersonnelMaster.objects.bulk_update(updated_personnel, ['current_rank', 'pending_rank'])
                AuditLog.objects.bulk_create(audit_logs)
                
        return Response({'success': True, 'message': f'تمت تسوية رتب {count} فرد بنجاح وتسجيل الحركات'})


# ═══════════════════════════════════════════════════
# RankSettlement CRUD + Apply/Reject
# ═══════════════════════════════════════════════════

class RankSettlementViewSet(PermissionRequiredMixin, BaseModelViewSet):
    """
    إدارة طلبات تسوية الرتب الرسمية.

    - GET    /rank-settlements/                  → قائمة الطلبات
    - POST   /rank-settlements/                  → إنشاء طلب جديد
    - GET    /rank-settlements/{id}/             → تفاصيل طلب
    - POST   /rank-settlements/{id}/apply/       → تطبيق (تغيير الرتبة فعلياً)
    - POST   /rank-settlements/{id}/reject/      → رفض
    """
    from .serializers import RankSettlementSerializer

    serializer_class = RankSettlementSerializer
    permission_map = {
        'list':     Perms.RANK_SETTLEMENT_VIEW,
        'retrieve': Perms.RANK_SETTLEMENT_VIEW,
        'create':   Perms.RANK_SETTLEMENT_CREATE,
        'apply':    Perms.RANK_SETTLEMENT_EXECUTE,
        'reject':   Perms.RANK_SETTLEMENT_EXECUTE,
    }
    
    filterset_fields = ['status', 'settlement_type']
    search_fields = [
        'personnel__military_number', 'personnel__full_name',
        'decision_number',
    ]
    ordering = ['-created_at']
    
    def get_queryset(self):
        from .models import RankSettlement
        return RankSettlement.objects.select_related(
            'personnel', 'from_rank', 'to_rank',
            'supporting_document', 'requested_by', 'applied_by',
        ).all()
    
    def perform_create(self, serializer):
        """حفظ الطالب + from_rank تلقائياً من الفرد"""
        personnel = serializer.validated_data['personnel']
        serializer.save(
            requested_by=self.request.user,
            from_rank=personnel.current_rank,
            status='pending',
        )
    
    @action(detail=True, methods=['post'])
    def apply(self, request, pk=None):
        """
        تطبيق طلب التسوية — يُفوّض لـ PersonnelService.apply_settlement
        يدعم: ترقية ضمن نفس الصنف، فرد→ضابط، تخفيض (عقوبة)
        """
        from .services import PersonnelService
        
        settlement = self.get_object()
        
        if settlement.status not in ('pending', 'approved'):
            return Response(
                {'success': False, 'error': 'الطلب ليس في حالة قابلة للتطبيق'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # الموافقة أولاً إذا كان pending
        if settlement.status == 'pending':
            settlement.status = 'approved'
            settlement.save(update_fields=['status'])
        
        ip_address = (
            request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
            or request.META.get('REMOTE_ADDR')
        )
        
        try:
            old_rank_name = settlement.from_rank.name
            PersonnelService.apply_settlement(
                settlement, user=request.user, ip_address=ip_address,
            )
            return Response({
                'success': True,
                'message': f'تم تطبيق التسوية: {old_rank_name} → {settlement.to_rank.name}',
                'settlement_type': settlement.settlement_type,
                'applied_by': request.user.username,
            })
        except Exception as e:
            return Response(
                {'success': False, 'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """رفض طلب تسوية"""
        from django.utils import timezone
        
        settlement = self.get_object()
        
        if settlement.status != 'pending':
            return Response(
                {'success': False, 'error': 'الطلب ليس في حالة الانتظار'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        reason = request.data.get('reason', 'تم الرفض')
        
        settlement.status = 'rejected'
        settlement.rejection_reason = reason
        settlement.applied_by = request.user
        settlement.applied_at = timezone.now()
        settlement.save(update_fields=['status', 'rejection_reason', 'applied_by', 'applied_at'])
        
        return Response({
            'success': True,
            'message': 'تم رفض طلب التسوية',
        })
