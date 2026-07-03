"""
أمر إدارة لترحيل البيانات التاريخية من ملفات Excel القديمة - المهمة 2.6

الاستخدام:
    python manage.py migrate_historical /path/to/old_file.xlsx

يقرأ ملف Excel قديم يحتوي على أعمدة "متغير شهر_سنة"
ويحولها إلى سجلات ServiceEventLog.
"""

import re
import logging
from datetime import date

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from systems.personnel.models import PersonnelMaster
from systems.services.models import ServiceEventLog

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'ترحيل البيانات التاريخية من ملفات Excel القديمة إلى ServiceEventLog'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='مسار ملف Excel القديم')
        parser.add_argument(
            '--dry-run', action='store_true', default=False,
            help='عرض النتائج بدون حفظ'
        )
        parser.add_argument(
            '--military-col', type=int, default=0,
            help='رقم عمود الرقم العسكري (يبدأ من 0)'
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        dry_run = options['dry_run']
        mil_col = options['military_col']

        self.stdout.write(f'قراءة الملف: {file_path}')

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f'فشل فتح الملف: {e}')

        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

        # اكتشاف أعمدة المتغيرات (regex: متغير + اسم_شهر + سنة)
        month_map = {
            'يناير': '01', 'فبراير': '02', 'مارس': '03', 'أبريل': '04',
            'مايو': '05', 'يونيو': '06', 'يوليو': '07', 'أغسطس': '08',
            'سبتمبر': '09', 'أكتوبر': '10', 'نوفمبر': '11', 'ديسمبر': '12',
            'كانون الثاني': '01', 'شباط': '02', 'آذار': '03', 'نيسان': '04',
            'أيار': '05', 'حزيران': '06', 'تموز': '07', 'آب': '08',
            'أيلول': '09', 'تشرين الأول': '10', 'تشرين الثاني': '11',
            'كانون الأول': '12',
        }

        variable_columns = []  # [(col_index, service_month)]

        for col_idx, header in enumerate(headers):
            if not header or col_idx == mil_col:
                continue

            # بحث عن نمط "متغير شهر سنة"
            for month_name, month_num in month_map.items():
                if month_name in header:
                    # استخراج السنة (4 أرقام)
                    year_match = re.search(r'(\d{4})', header)
                    if year_match:
                        year = year_match.group(1)
                        service_month = f'{year}-{month_num}'
                        variable_columns.append((col_idx, service_month))
                        break

        if not variable_columns:
            raise CommandError('لم يتم اكتشاف أي أعمدة متغيرات شهرية')

        self.stdout.write(
            f'تم اكتشاف {len(variable_columns)} عمود متغير: '
            + ', '.join([sm for _, sm in variable_columns])
        )

        # معالجة الصفوف
        stats = {
            'total_rows': 0,
            'events_created': 0,
            'not_found': 0,
            'errors': 0,
            'skipped_empty': 0,
        }
        errors_log = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[mil_col]:
                continue

            stats['total_rows'] += 1
            military_number = str(row[mil_col]).strip()

            # البحث عن الفرد
            try:
                personnel = PersonnelMaster.objects.get(
                    military_number=military_number
                )
            except PersonnelMaster.DoesNotExist:
                stats['not_found'] += 1
                errors_log.append(f'صف {row_num}: رقم عسكري غير موجود: {military_number}')
                continue

            # لكل عمود متغير
            previous_status = ''
            for col_idx, service_month in variable_columns:
                if col_idx >= len(row) or not row[col_idx]:
                    stats['skipped_empty'] += 1
                    continue

                new_value = str(row[col_idx]).strip()
                if not new_value:
                    stats['skipped_empty'] += 1
                    continue

                # إنشاء ServiceEventLog
                try:
                    year, month = service_month.split('-')
                    event_date = date(int(year), int(month), 1)

                    if not dry_run:
                        ServiceEventLog.objects.update_or_create(
                            personnel=personnel,
                            service_month=service_month,
                            field_name='current_status',
                            defaults={
                                'event_date': event_date,
                                'old_value': previous_status,
                                'new_value': new_value,
                            }
                        )

                    stats['events_created'] += 1
                    previous_status = new_value

                except Exception as e:
                    stats['errors'] += 1
                    errors_log.append(
                        f'صف {row_num}, شهر {service_month}: {str(e)}'
                    )

        wb.close()

        # عرض النتائج
        self.stdout.write(self.style.SUCCESS(f'\n=== نتائج الترحيل ==='))
        self.stdout.write(f'إجمالي الصفوف: {stats["total_rows"]}')
        self.stdout.write(f'أحداث تم إنشاؤها: {stats["events_created"]}')
        self.stdout.write(f'أرقام عسكرية غير موجودة: {stats["not_found"]}')
        self.stdout.write(f'خلايا فارغة (تم تخطيها): {stats["skipped_empty"]}')
        self.stdout.write(f'أخطاء: {stats["errors"]}')

        if dry_run:
            self.stdout.write(self.style.WARNING('\n⚠️ هذا تشغيل تجريبي (dry-run). لم يتم حفظ أي بيانات.'))

        if errors_log:
            self.stdout.write(self.style.ERROR(f'\n=== الأخطاء ({len(errors_log)}) ==='))
            for err in errors_log[:20]:
                self.stdout.write(f'  - {err}')
            if len(errors_log) > 20:
                self.stdout.write(f'  ... و {len(errors_log) - 20} خطأ آخر')
