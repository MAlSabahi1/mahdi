"""
اختبارات خدمة التصدير - المهمة 2.1

تتحقق من:
- إنشاء ملف Excel مع 4 أوراق عمل
- الأعمدة الصحيحة (8 أعمدة)
- حماية المصنف والورقة
- UUID مخفي لكل صف  
- حساب Hash
- تسجيل ExportLog
"""
import pytest
from io import BytesIO
from django.contrib.auth import get_user_model
from django.utils import timezone
import openpyxl

from core.models import GeoGovernorate, SecurityAdministration, CentralDepartment, GeoDistrict, Rank, ServiceStatus, Qualification
from systems.personnel.models import PersonnelMaster
from systems.services.export_service import ExcelExportService, export_template_for_department
from systems.services.models import ExportLog

User = get_user_model()


@pytest.fixture
def setup_test_data(db):
    """تجهيز بيانات الاختبار"""
    # إنشاء مستخدم
    user = User.objects.create_user(
        username='testuser',
        password='testpass123',
        is_staff=True
    )
    
    # إنشاء محافظة
    governorate = GeoGovernorate.objects.create(name_ar='بغداد')
    security_admin = SecurityAdministration.objects.create(name='أمن العاصمة', geo_governorate=governorate)

    # إنشاء إدارة
    department = CentralDepartment.objects.create(
        name='إدارة الاختبار',
        code='TEST001',
       
        security_admin=security_admin
    )
    
    # إنشاء رتبة
    rank = Rank.objects.create(
        name='جندي',
        order=1,
        is_officer=False
    )
    
    # إنشاء حالات خدمية
    status_active = ServiceStatus.objects.create(
        name='عامل',
        classification='active_full',
        receives_salary=True,
        requires_document=False
    )
    
    status_inactive = ServiceStatus.objects.create(
        name='منتدب',
        classification='inactive_temp',
        receives_salary=True,
        requires_document=True
    )
    
    status_absent = ServiceStatus.objects.create(
        name='غياب',
        classification='inactive_temp',
        receives_salary=False,
        requires_document=False
    )
    
    status_perm = ServiceStatus.objects.create(
        name='متوفى',
        classification='inactive_perm',
        receives_salary=False,
        requires_document=True
    )
    
    # إنشاء مؤهل
    qualification = Qualification.objects.create(
        name='ثانوية',
        order=1
    )
    
    # إنشاء موقع
    location = GeoDistrict.objects.create(
        name='بغداد',
        code='BGD'
    )
    
    # إنشاء أفراد (مع حالات مختلفة)
    personnel_list = []
    statuses = [status_active, status_active, status_active, status_inactive, status_absent]
    
    for i in range(5):
        person = PersonnelMaster.objects.create(
            military_number=f'100000{i}',
            national_id=f'1234567890{i}',
            full_name=f'اسم الاختبار {i}',
            birth_date='1990-01-01',
            join_date='2010-01-01',
            qualification=qualification,
            geo_location=location,
            central_department=department,
            current_rank=rank,
            current_status=statuses[i]
        )
        personnel_list.append(person)
    
    return {
        'user': user,
        'central_department': department,
        'rank': rank,
        'status_active': status_active,
        'status_inactive': status_inactive,
        'status_absent': status_absent,
        'status_perm': status_perm,
        'qualification': qualification,
        'location': location,
        'personnel': personnel_list
    }


@pytest.mark.django_db
class TestExcelExportService:
    """اختبارات خدمة التصدير"""
    
    def test_service_initialization(self, setup_test_data):
        """اختبار تهيئة الخدمة"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        assert service.central_department == data['central_department']
        assert service.service_month == '2026-03'
        assert service.exported_by == data['user']
        assert service.row_uuids == []
        assert service.file_hash is None
    
    def test_get_personnel_data(self, setup_test_data):
        """اختبار استخراج بيانات الأفراد"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        personnel_data = service.get_personnel_data()
        
        assert len(personnel_data) == 5
        assert len(service.row_uuids) == 5
        
        # التحقق من وجود جميع الحقول المطلوبة
        first_person = personnel_data[0]
        assert 'military_number' in first_person
        assert 'full_name' in first_person
        assert 'rank' in first_person
        assert 'national_id' in first_person
        assert 'current_status' in first_person
        assert 'classification' in first_person
        assert 'row_uuid' in first_person
        
        # التحقق من أن UUIDs فريدة
        uuids = [p['row_uuid'] for p in personnel_data]
        assert len(uuids) == len(set(uuids))
    
    def test_classify_personnel(self, setup_test_data):
        """اختبار توزيع الأفراد على الأوراق الأربعة"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        personnel_data = service.get_personnel_data()
        sheets = service._classify_personnel(personnel_data)
        
        # 3 عاملين
        assert len(sheets['القوة العاملة']) == 3
        # 1 منتدب + 1 غياب = 2 غير عاملين
        assert len(sheets['القوة غير العاملة']) == 2
        # الجميع
        assert len(sheets['القوة كاملة']) == 5
        # 1 غياب
        assert len(sheets['الغياب']) == 1
    
    def test_get_allowed_statuses(self, setup_test_data):
        """اختبار الحصول على الحالات المسموحة"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        statuses = service.get_allowed_statuses()
        
        assert len(statuses) > 0
        assert 'عامل' in statuses
        assert 'أخرى' in statuses
    
    def test_create_protected_excel_has_4_sheets(self, setup_test_data):
        """اختبار أن ملف Excel يحتوي على 4 أوراق"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        excel_file = service.create_protected_excel()
        
        # فتح الملف للتحقق
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        assert len(wb.sheetnames) == 4
        assert 'القوة العاملة' in wb.sheetnames
        assert 'القوة غير العاملة' in wb.sheetnames
        assert 'القوة كاملة' in wb.sheetnames
        assert 'الغياب' in wb.sheetnames
        
        wb.close()
    
    def test_create_protected_excel_correct_columns(self, setup_test_data):
        """اختبار أن الأعمدة صحيحة (8 أعمدة)"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        excel_file = service.create_protected_excel()
        
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb['القوة كاملة']
        
        # قراءة العناوين
        headers = [cell.value for cell in ws[1]]
        
        assert headers[0] == 'الرقم العسكري'
        assert headers[1] == 'الاسم الكامل'
        assert headers[2] == 'الرتبة'
        assert headers[3] == 'الرقم الوطني'
        assert headers[4] == 'الحالة الحالية'
        assert headers[5] == 'متغير الشهر'
        assert headers[6] == 'ملاحظات'
        assert headers[7] == '__UUID__'
        
        wb.close()
    
    def test_create_protected_excel_hash(self, setup_test_data):
        """اختبار حساب Hash"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        excel_file = service.create_protected_excel()
        
        assert isinstance(excel_file, BytesIO)
        assert excel_file.getvalue()  # الملف ليس فارغاً
        assert service.file_hash is not None
        assert len(service.file_hash) == 64  # SHA-256 hex length
    
    def test_export_and_log(self, setup_test_data):
        """اختبار التصدير والتسجيل في قاعدة البيانات"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        excel_file, filename = service.export_and_log()
        
        # التحقق من الملف
        assert isinstance(excel_file, BytesIO)
        assert filename.endswith('.xlsx')
        assert 'إدارة الاختبار' in filename
        assert '2026-03' in filename
        
        # التحقق من التسجيل في قاعدة البيانات
        export_log = ExportLog.objects.latest('created_at')
        assert export_log.central_department == data['central_department']
        assert export_log.service_month == '2026-03'
        assert export_log.exported_by == data['user']
        assert export_log.file_hash == service.file_hash
        assert len(export_log.row_uuids) == 5
        assert export_log.status == 'pending'
    
    def test_export_template_for_department(self, setup_test_data):
        """اختبار الدالة المساعدة للتصدير"""
        data = setup_test_data
        
        excel_file, filename = export_template_for_department(
            central_department_id=data['central_department'].id,
            service_month='2026-03',
            user=data['user']
        )
        
        assert isinstance(excel_file, BytesIO)
        assert filename.endswith('.xlsx')
        
        # التحقق من وجود السجل
        assert ExportLog.objects.filter(
            central_department=data['central_department'],
            service_month='2026-03'
        ).exists()
    
    def test_export_empty_department(self, setup_test_data):
        """اختبار تصدير إدارة بدون أفراد"""
        data = setup_test_data
        
        # إنشاء إدارة فارغة
        empty_dept = CentralDepartment.objects.create(
            name='إدارة فارغة',
            code='EMPTY',
           
            governorate=data['central_department'].governorate
        )
        
        service = ExcelExportService(
            central_department=empty_dept,
            service_month='2026-03',
            exported_by=data['user']
        )
        
        personnel_data = service.get_personnel_data()
        assert len(personnel_data) == 0
        
        # يجب أن يعمل التصدير حتى للإدارة الفارغة
        excel_file = service.create_protected_excel()
        assert isinstance(excel_file, BytesIO)
        
        # التحقق من 4 أوراق حتى للإدارة الفارغة
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        assert len(wb.sheetnames) == 4
        wb.close()
    
    def test_password_generation(self, setup_test_data):
        """اختبار توليد كلمة المرور"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        password1 = service._generate_password()
        password2 = service._generate_password()
        
        # نفس المعاملات تعطي نفس كلمة المرور
        assert password1 == password2
        assert len(password1) == 16
    
    def test_hash_calculation(self, setup_test_data):
        """اختبار حساب Hash"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        test_content = b"test content"
        hash1 = service._calculate_hash(test_content)
        hash2 = service._calculate_hash(test_content)
        
        assert hash1 == hash2
        assert len(hash1) == 64  # SHA-256 hex
    
    def test_multiple_exports_same_central_department(self, setup_test_data):
        """اختبار تصديرات متعددة لنفس الإدارة"""
        data = setup_test_data
        
        # التصدير الأول
        excel1, filename1 = export_template_for_department(
            central_department_id=data['central_department'].id,
            service_month='2026-03',
            user=data['user']
        )
        
        # التصدير الثاني
        excel2, filename2 = export_template_for_department(
            central_department_id=data['central_department'].id,
            service_month='2026-04',
            user=data['user']
        )
        
        # يجب أن يكون لكل تصدير export_id مختلف
        exports = ExportLog.objects.filter(central_department=data['central_department'])
        assert exports.count() == 2
        
        export_ids = [e.export_id for e in exports]
        assert len(export_ids) == len(set(export_ids))  # جميع IDs فريدة
    
    def test_correct_data_in_sheets(self, setup_test_data):
        """اختبار صحة البيانات في كل ورقة"""
        data = setup_test_data
        service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        
        excel_file = service.create_protected_excel()
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # ورقة القوة كاملة يجب أن تحتوي على 5 أفراد
        ws_all = wb['القوة كاملة']
        rows = list(ws_all.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if r[0]]
        assert len(non_empty) == 5
        
        # ورقة القوة العاملة يجب أن تحتوي على 3 أفراد
        ws_active = wb['القوة العاملة']
        rows_active = list(ws_active.iter_rows(min_row=2, values_only=True))
        non_empty_active = [r for r in rows_active if r[0]]
        assert len(non_empty_active) == 3
        
        wb.close()
