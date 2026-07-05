"""
Import Service - خدمة استيراد الكشوفات المعدلة (إنتاجي)
═══════════════════════════════════════════════════════════
سلسلة التحقق الكاملة:
1. التحقق من اسم الملف (مطابقة النمط الرسمي)
2. التحقق من ExportLog (هل صادر منا؟)
3. التحقق من بنية الأوراق (single أو multi)
4. فحص الماكروز والمحتوى الخطر
5. التحقق من العدد الإجمالي (لا زيادة ولا نقصان)
6. لكل صف: UUID + الرقم العسكري + الاسم + الرقم الوطني
7. كشف التلاعب بالأعمدة المقفولة
8. اكتشاف التغييرات في الأعمدة المسموحة فقط
9. تصنيف كل تغيير: أخضر / أصفر / أحمر + تحديد المرفقات المطلوبة
10. وضع التغييرات في StagingRecord
"""
import hashlib
import uuid
import re
import zipfile
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from io import BytesIO

import openpyxl
from django.db import transaction
from django.utils import timezone
from django.core.exceptions import ValidationError

from systems.personnel.models import PersonnelMaster
from core.models import ServiceStatus, CentralDepartment
from systems.services.models import ExportLog, StagingRecord, AuditLog

# المرفقات المطلوبة لكل حالة حساسة
REQUIRED_ATTACHMENTS: Dict[str, List[str]] = {
    'الشهداء':               ['شهادة استشهاد', 'تقرير الواقعة', 'قرار الاعتراف'],
    'الوفيات':               ['شهادة وفاة', 'شهادة الطبيب الشرعي'],
    'المتقاعدين':            ['قرار التقاعد', 'موافقة الوزارة'],
    'المفقودين':             ['صك شرعي', 'تقرير اللجنة', 'موافقة وزارية'],
    'المفقودين (بصك شرعي)': ['صك شرعي موثق', 'تقرير اللجنة'],
    'المنقطعين - فرار':      ['بلاغ انقطاع', 'قرار تشكيل لجنة'],
    'المفرغين للدراسة':      ['قرار الابتعاث', 'موافقة الجهة'],
    'المنتدبين لدى جهات':   ['قرار الندب', 'موافقة المحافظة'],
    'السجناء':               ['حكم قضائي أو قرار توقيف'],
    'الأسرى':                ['تقرير رسمي'],
    'عدم لياقة':             ['تقرير اللجنة الطبية', 'موافقة وزارية'],
    'إنهاء المدة القانونية': ['قرار إنهاء الخدمة'],
    'بلوغ السن القانوني':   ['موافقة وزارية', 'قرار التقاعد'],
}

# الأعمدة المحمية التي يُعدّ تعديلها تلاعباً
PROTECTED_FIELD_MAP = {
    'الرقم العسكري':  'military_number',
    'الاسم الكامل':   'full_name',
    'الرتبة':         'current_rank__name',
    'الرقم الوطني':   'national_id',
    'الحالة الحالية': 'current_status__name',
}

# نمط اسم الملف الرسمي
FILENAME_PATTERN = re.compile(
    r'^كشف_(?P<dept>.+)_(?P<month>\d{4}-\d{2})_(?P<uid>[a-f0-9]{8})\.xlsx$'
)

# كلمات تدل على ماكروز أو محتوى خطر
DANGEROUS_PATTERNS = [b'xl/vbaProject', b'<vba', b'Auto_Open', b'Workbook_Open']


class ImportValidationError(Exception):
    """فشل التحقق — يوقف المعالجة بالكامل"""
    pass


class ExcelImportService:
    """
    خدمة استيراد الكشوفات المعدلة مع تحقق أمني متكامل.
    """

    EXPECTED_SHEETS_MULTI  = ['القوة العاملة', 'القوة غير العاملة', 'القوة كاملة', 'الغياب']
    EXPECTED_SHEETS_SINGLE = ['كشف موحد']

    def __init__(self, file_content: bytes, export_id: str, imported_by, original_filename: str = '', service_month: str = ''):
        self.file_content      = file_content
        self.export_id         = export_id
        self.imported_by       = imported_by
        self.original_filename = original_filename.strip()
        self.service_month     = service_month

        self.export_log  = None
        self.workbook    = None
        self.errors:   List[Dict] = []
        self.warnings: List[Dict] = []
        self.changes:  List[Dict] = []
        self.batch_id = uuid.uuid4()

        self.stats = {
            'total_rows': 0, 'changes_detected': 0,
            'errors': 0, 'warnings': 0,
            'green': 0, 'yellow': 0, 'red': 0,
        }

    # ─────────────────────── خطوات التحقق ───────────────────────

    def _step1_verify_filename(self):
        """الخطوة 1: التحقق من اسم الملف."""
        if not self.original_filename:
            return  # اسم الملف اختياري في بعض الاستخدامات
        m = FILENAME_PATTERN.match(self.original_filename)
        if not m:
            raise ImportValidationError(
                f"اسم الملف غير مطابق للصيغة الرسمية. "
                f"الصيغة المطلوبة: كشف_[الإدارة]_[YYYY-MM]_[uid].xlsx\n"
                f"المُستلم: {self.original_filename}"
            )
        # التحقق من الشهر المُضمَّن في الاسم
        if self.service_month and m.group('month') != self.service_month:
            raise ImportValidationError(
                f"الشهر في اسم الملف ({m.group('month')}) "
                f"لا يطابق الشهر المُمرَّر ({self.service_month})"
            )

    def _step2_verify_export_log(self):
        """الخطوة 2: التحقق من وجود سجل التصدير."""
        try:
            self.export_log = ExportLog.objects.select_related(
                'central_department', 'security_admin'
            ).get(export_id=self.export_id)
        except ExportLog.DoesNotExist:
            raise ImportValidationError(f"لم يُعثر على سجل تصدير بالمعرف: {self.export_id}")

        if self.export_log.status == 'expired':
            raise ImportValidationError("هذا الملف منتهي الصلاحية.")

        if self.export_log.status == 'returned':
            self.warnings.append({'warning': 'هذا الملف سبق استيراده. سيتم معالجته مجدداً.'})

    def _step3_scan_for_malware(self):
        """الخطوة 3: فحص المحتوى الخطر (ماكروز - فيروسات)."""
        for pattern in DANGEROUS_PATTERNS:
            if pattern in self.file_content:
                raise ImportValidationError(
                    f"تم رفض الملف: يحتوي على محتوى خطر ({pattern.decode(errors='ignore')}). "
                    "لا يُسمح بالماكروز أو الأكواد القابلة للتنفيذ."
                )

    def _step4_load_workbook(self):
        """الخطوة 4: تحميل الملف والتحقق من بنيته."""
        try:
            self.workbook = openpyxl.load_workbook(
                BytesIO(self.file_content), data_only=True, read_only=False
            )
        except Exception as e:
            raise ImportValidationError(f"فشل قراءة ملف Excel: {e}")

        sheet_names = self.workbook.sheetnames

        # تحديد الوضع (single أو multi)
        if 'كشف موحد' in sheet_names:
            self.sheet_mode = 'single'
            self.expected_sheets = self.EXPECTED_SHEETS_SINGLE
        else:
            self.sheet_mode = 'multi'
            self.expected_sheets = self.EXPECTED_SHEETS_MULTI

        for expected in self.expected_sheets:
            if expected not in sheet_names:
                raise ImportValidationError(f"الورقة '{expected}' مفقودة.")

        # التحقق من عدم إضافة أوراق إضافية
        extra = set(sheet_names) - set(self.expected_sheets)
        if extra:
            raise ImportValidationError(f"تم اكتشاف أوراق إضافية غير مصرح بها: {extra}")

    def _step5_verify_row_count(self, all_rows: List[Dict]):
        """الخطوة 5: التحقق من عدد الصفوف (لا زيادة ولا نقصان)."""
        expected_count = len(self.export_log.row_uuids)

        # حساب الصفوف الفعلية (بدون تكرار UUID في أوراق متعددة)
        seen_uuids = set()
        unique_count = 0
        for row in all_rows:
            uid = str(row.get('__UUID__', ''))
            if uid and uid not in seen_uuids:
                seen_uuids.add(uid)
                unique_count += 1

        if unique_count != expected_count:
            raise ImportValidationError(
                f"عدد الصفوف غير متطابق! "
                f"المُصدَّر: {expected_count} فرد، المُستلَم: {unique_count} فرد. "
                "يُمنع إضافة أو حذف أي صف."
            )

    def _read_all_rows(self) -> List[Dict]:
        """قراءة كل الصفوف من كل الأوراق."""
        all_rows = []
        for sheet_name in self.expected_sheets:
            ws = self.workbook[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            # إيجاد صف العناوين (تجاوز صف التحذير الأول)
            # صف 0: التحذير، صف 1: العناوين
            headers = [str(h).strip() if h else '' for h in rows[1]]

            for row_idx, row_vals in enumerate(rows[2:], start=3):
                if not any(row_vals):
                    continue
                row_dict = {}
                for col_i, h in enumerate(headers):
                    if col_i < len(row_vals):
                        val = row_vals[col_i]
                        row_dict[h] = str(val).strip() if val is not None else ''
                row_dict['_sheet'] = sheet_name
                row_dict['_row']   = row_idx
                all_rows.append(row_dict)

        return all_rows

    def _step6_detect_changes(self, all_rows: List[Dict]):
        """الخطوة 6–9: كشف التلاعب والتغييرات وتصنيفها."""
        exported_uuids = {str(u) for u in self.export_log.row_uuids}
        editable_cols  = set(self.export_log.editable_columns)
        processed_mil  = set()

        for row in all_rows:
            self.stats['total_rows'] += 1
            sheet = row.get('_sheet', '')
            row_n = row.get('_row', 0)
            mil_num = row.get('الرقم العسكري', '').strip()
            row_uuid = row.get('__UUID__', '').strip()

            # ── التحقق من UUID ──
            if row_uuid not in exported_uuids:
                self._add_error(sheet, row_n, mil_num, 'UUID غير صالح — صف غير أصلي أو مُزوَّر')
                continue

            if mil_num in processed_mil:
                continue
            processed_mil.add(mil_num)

            # ── البحث في DB ──
            try:
                person = PersonnelMaster.objects.select_related(
                    'current_rank', 'current_status'
                ).get(military_number=mil_num)
            except PersonnelMaster.DoesNotExist:
                self._add_error(sheet, row_n, mil_num, 'الرقم العسكري غير موجود في قاعدة البيانات')
                continue

            # ── الخطوة 7: كشف التلاعب بالأعمدة المقفولة ──
            tampering = self._check_tampering(row, person)
            for t in tampering:
                self.warnings.append({'sheet': sheet, 'row': row_n, 'military_number': mil_num, **t})
                self.stats['warnings'] += 1

            # ── الخطوة 8: كشف التغييرات المسموح بها ──
            new_status_cls  = row.get('الحالة الجديدة', '').strip()
            new_status_type = row.get('نوع الحالة الجديدة', '').strip()
            notes           = row.get('ملاحظات', '').strip()
            old_status      = person.current_status.name if person.current_status else ''

            if new_status_type and new_status_type != old_status:
                severity, requires_doc, attachments = self._classify_change(new_status_type)
                self.stats['changes_detected'] += 1
                self.stats[severity] += 1

                self.changes.append({
                    'personnel':         person,
                    'old_status':        old_status,
                    'new_status':        new_status_type,
                    'new_classification': new_status_cls,
                    'notes':             notes,
                    'severity':          severity,
                    'requires_document': requires_doc,
                    'required_attachments': attachments,
                    'tampered_fields':   tampering,
                    'sheet':             sheet,
                    'row':               row_n,
                })

    def _check_tampering(self, row: Dict, person: PersonnelMaster) -> List[Dict]:
        """فحص التلاعب بالأعمدة المقفولة."""
        issues = []
        db_vals = {
            'الرقم العسكري':  person.military_number or '',
            'الاسم الكامل':   person.full_name or '',
            'الرتبة':         person.current_rank.name if person.current_rank else '',
            'الرقم الوطني':   person.national_id or '',
            'الحالة الحالية': person.current_status.name if person.current_status else '',
        }
        for col, db_val in db_vals.items():
            file_val = row.get(col, '').strip()
            if file_val and file_val != str(db_val).strip():
                issues.append({
                    'type':      'tampered_locked_column',
                    'column':    col,
                    'file_value': file_val,
                    'db_value':  db_val,
                    'warning':   f'تلاعب في العمود المقفول "{col}": الملف="{file_val}" | DB="{db_val}"',
                })
        return issues

    def _classify_change(self, new_status: str) -> Tuple[str, bool, List[str]]:
        """
        تصنيف التغيير:
        Returns: (severity, requires_document, required_attachments)
        - green: لا يحتاج مستند
        - yellow: يحتاج مستند محدد
        - red: مريب أو غير معروف
        """
        attachments = REQUIRED_ATTACHMENTS.get(new_status, [])
        if attachments:
            return 'yellow', True, attachments
        try:
            ServiceStatus.objects.get(name=new_status)
            return 'green', False, []
        except ServiceStatus.DoesNotExist:
            return 'red', True, ['وثيقة تثبت صحة الحالة']

    def _add_error(self, sheet, row, mil, msg):
        self.errors.append({'sheet': sheet, 'row': row, 'military_number': mil, 'error': msg})
        self.stats['errors'] += 1

    @transaction.atomic
    def _create_staging_records(self):
        """إنشاء سجلات StagingRecord دفعة واحدة."""
        records = []
        for ch in self.changes:
            records.append(StagingRecord(
                personnel=ch['personnel'],
                security_admin=self.export_log.security_admin,
                upload_batch_id=self.batch_id,
                proposed_change={
                    'status':              ch['new_status'],
                    'classification':      ch['new_classification'],
                    'service_month':       self.service_month or self.export_log.service_month,
                    'required_attachments': ch['required_attachments'],
                },
                notes=ch['notes'],
                status='pending',
                severity=ch['severity'],
                requires_document=ch['requires_document'],
                name_mismatch=any(t['column'] == 'الاسم الكامل' for t in ch['tampered_fields']),
                rank_mismatch=any(t['column'] == 'الرتبة' for t in ch['tampered_fields']),
                national_id_mismatch=any(t['column'] == 'الرقم الوطني' for t in ch['tampered_fields']),
            ))

        StagingRecord.objects.bulk_create(records)

        self.export_log.status = 'returned'
        self.export_log.save(update_fields=['status'])

        AuditLog.objects.create(
            user=self.imported_by,
            action='IMPORT',
            model_name='ExportLog',
            object_id=str(self.export_log.export_id),
            new_data={
                'batch_id':      str(self.batch_id),
                'changes':       self.stats['changes_detected'],
                'errors':        self.stats['errors'],
                'green':         self.stats['green'],
                'yellow':        self.stats['yellow'],
                'red':           self.stats['red'],
                'filename':      self.original_filename,
            },
        )

    @transaction.atomic
    def process(self) -> Dict[str, Any]:
        """تشغيل سلسلة التحقق الكاملة."""
        self._step1_verify_filename()
        self._step2_verify_export_log()
        self._step3_scan_for_malware()
        self._step4_load_workbook()

        all_rows = self._read_all_rows()

        self._step5_verify_row_count(all_rows)
        self._step6_detect_changes(all_rows)

        if self.changes:
            self._create_staging_records()

        # دائماً نسجّل حالة الإرجاع + AuditLog حتى بدون تغييرات
        self.export_log.status = 'returned'
        self.export_log.save(update_fields=['status'])

        AuditLog.objects.create(
            user=self.imported_by,
            action='IMPORT',
            model_name='ExportLog',
            object_id=str(self.export_log.export_id),
            new_data={
                'batch_id': str(self.batch_id),
                'changes':  self.stats['changes_detected'],
                'errors':   self.stats['errors'],
                'green':    self.stats['green'],
                'yellow':   self.stats['yellow'],
                'red':      self.stats['red'],
                'filename': self.original_filename,
            },
        )

        return {
            'success':    True,
            'batch_id':   str(self.batch_id),
            'department': self.export_log.central_department.name if self.export_log.central_department else '—',
            'month':      self.service_month or self.export_log.service_month,
            'stats':      self.stats,
            'errors':     self.errors,
            'warnings':   self.warnings,
            'summary':    self._summary(),
        }

    def _summary(self) -> str:
        parts = [f"إجمالي الصفوف: {self.stats['total_rows']}"]
        if self.stats['changes_detected']:
            parts.append(
                f"تغييرات: {self.stats['changes_detected']} "
                f"(🟢{self.stats['green']} 🟡{self.stats['yellow']} 🔴{self.stats['red']})"
            )
        if self.stats['errors']:
            parts.append(f"أخطاء: {self.stats['errors']}")
        if self.stats['warnings']:
            parts.append(f"تحذيرات: {self.stats['warnings']}")
        return ' | '.join(parts)


def import_service_file(file_content: bytes, export_id: str, user, service_month: str = '', filename: str = '') -> Dict:
    svc = ExcelImportService(
        file_content=file_content,
        export_id=export_id,
        imported_by=user,
        original_filename=filename,
        service_month=service_month,
    )
    return svc.process()
