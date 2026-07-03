"""
خدمة المراجعة والاعتماد - جسر (Bridge)
تم نقل المنطق إلى أنظمة Clean Architecture (Use Cases).
هذا الملف يعمل كواجهة توافقية (Backward Compatibility).
"""

from typing import Dict, Optional
from infra.storage.models import Document
from systems.services.infrastructure.models.staging import StagingRecord

# Import Clean Architecture Components
from systems.services.domain.entities.staging import StagingRecordEntity
from systems.services.infrastructure.repositories.django_staging_repo import DjangoStagingRecordRepository
from systems.personnel.infrastructure.repositories.django_personnel_repo import DjangoPersonnelRepository
from systems.services.application.use_cases.staging_use_cases import (
    ReviewStagingCommand,
    ReviewStagingRecordUseCase,
    AutoApproveBatchCommand,
    AutoApproveLowSeverityUseCase
)


class ReviewValidationError(Exception):
    pass


class StagingReviewService:
    def __init__(self, user):
        self.user = user
        self.repo = DjangoStagingRecordRepository()
        self.personnel_repo = DjangoPersonnelRepository()

    def get_pending_records(self, batch_id: str = None, central_department_id: int = None) -> Dict:
        qs = StagingRecord.objects.filter(status='pending').select_related(
            'personnel', 'personnel__current_rank', 
            'personnel__current_status', 'personnel__central_department'
        )
        if batch_id:
            qs = qs.filter(upload_batch_id=batch_id)
        if central_department_id:
            qs = qs.filter(personnel__central_department_id=central_department_id)
            
        records = list(qs)
        low_severity = [r for r in records if r.severity == 'low']
        high_severity = [r for r in records if r.severity == 'high']
        with_mismatches = [r for r in records if r.name_mismatch or r.rank_mismatch or r.national_id_mismatch]
        
        return {
            'low_severity': low_severity,
            'high_severity': high_severity,
            'with_mismatches': with_mismatches,
            'total': len(records),
            'all': records
        }

    def approve_record(self, staging_id: int, document: Optional[Document] = None) -> Dict:
        cmd = ReviewStagingCommand(
            record_id=staging_id,
            reviewer_id=self.user.id,
            action='approve',
            document_id=document.id if document else None
        )
        uc = ReviewStagingRecordUseCase(self.repo, self.personnel_repo)
        
        try:
            result = uc.execute(cmd)
            # Fetch personnel to return required dict
            personnel = self.personnel_repo.get_by_military_number(result.military_number)
            return {
                'success': True,
                'military_number': result.military_number,
                'new_status': result.changes.get('status', {}).get('new', '')
            }
        except ValueError as e:
            raise ReviewValidationError(str(e))

    def reject_record(self, staging_id: int, reason: str) -> Dict:
        cmd = ReviewStagingCommand(
            record_id=staging_id,
            reviewer_id=self.user.id,
            action='reject',
            reason=reason
        )
        uc = ReviewStagingRecordUseCase(self.repo, self.personnel_repo)
        
        try:
            result = uc.execute(cmd)
            return {
                'success': True,
                'military_number': result.military_number,
                'rejection_reason': reason
            }
        except ValueError as e:
            raise ReviewValidationError(str(e))

    def approve_all_low_severity(self, batch_id: str = None, central_department_id: int = None) -> Dict:
        cmd = AutoApproveBatchCommand(
            batch_id=batch_id or '',
            reviewer_id=self.user.id
        )
        uc = AutoApproveLowSeverityUseCase(self.repo, self.personnel_repo)
        
        try:
            approved = uc.execute(cmd)
            return {
                'success': True,
                'approved_count': len(approved),
                'error_count': 0,
                'errors': []
            }
        except Exception as e:
            return {
                'success': False,
                'approved_count': 0,
                'error_count': 1,
                'errors': [{'error': str(e)}]
            }

    def reopen_rejection(self, rejection_id: int) -> Dict:
        # For simplicity, reopen rejection can remain directly via ORM in bridge
        from systems.services.infrastructure.models.logs import RejectionLog
        from infra.audit.models import AuditLog
        try:
            rejection = RejectionLog.objects.select_related('staging_record', 'personnel').get(id=rejection_id)
        except RejectionLog.DoesNotExist:
            raise ReviewValidationError(f'سجل الرفض غير موجود: {rejection_id}')
        
        old_record = rejection.staging_record
        new_record = StagingRecord.objects.create(
            personnel=old_record.personnel,
            upload_batch_id=old_record.upload_batch_id,
            proposed_change=old_record.proposed_change,
            notes=old_record.notes,
            status='pending',
            severity=old_record.severity,
            requires_document=old_record.requires_document,
            name_mismatch=old_record.name_mismatch,
            rank_mismatch=old_record.rank_mismatch,
            national_id_mismatch=old_record.national_id_mismatch
        )
        AuditLog.objects.create(
            user=self.user,
            action='CREATE',
            model_name='StagingRecord',
            object_id=str(new_record.id),
            old_data={'rejection_id': rejection_id, 'action': 'reopen'},
            new_data={'status': 'pending'}
        )
        return {
            'success': True,
            'new_staging_id': new_record.id,
            'military_number': old_record.personnel.military_number
        }

    def export_rejection_report(self, central_department_id: int, service_month: str) -> tuple:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from systems.services.infrastructure.models.logs import RejectionLog
        
        rejections = RejectionLog.objects.filter(
            central_department_id=central_department_id,
            service_month=service_month
        ).select_related(
            'personnel', 'personnel__current_rank',
            'rejected_by', 'central_department'
        ).order_by('-created_at')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'التغييرات المرفوضة'
        
        headers = ['الرقم العسكري', 'الاسم الكامل', 'الرتبة', 'الحالة المقترحة', 'سبب الرفض', 'رفض بواسطة', 'تاريخ الرفض']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, rej in enumerate(rejections, 2):
            personnel = rej.personnel
            ws.cell(row=row_idx, column=1, value=personnel.military_number)
            ws.cell(row=row_idx, column=2, value=personnel.full_name)
            ws.cell(row=row_idx, column=3, value=personnel.current_rank.name if personnel.current_rank else '')
            ws.cell(row=row_idx, column=4, value=rej.proposed_status)
            ws.cell(row=row_idx, column=5, value=rej.rejection_reason)
            ws.cell(row=row_idx, column=6, value=rej.rejected_by.username if rej.rejected_by else '')
            ws.cell(row=row_idx, column=7, value=rej.created_at.strftime('%Y-%m-%d %H:%M') if rej.created_at else '')
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        dept_name = rejections[0].central_department.name if rejections.exists() and rejections[0].central_department else 'إدارة'
        filename = f'تقرير_المرفوضات_{dept_name}_{service_month}.xlsx'
        
        return output, filename
