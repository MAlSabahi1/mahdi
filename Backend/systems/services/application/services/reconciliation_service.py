"""
محرك المطابقة - المهمة 2.4

المسؤوليات:
1. مقارنة ملف Excel خارجي مع قاعدة البيانات
2. تصنيف النتائج إلى 4 فئات: matched, different, new_in_file, missing_from_file
3. دعم 3 أنواع مطابقة: attendance, payroll, qualification
4. Fuzzy Matching بالاسم (Levenshtein ≥ 85%)
5. واجهة لتطبيق التغييرات المكتشفة
"""

import uuid
from datetime import date
from difflib import SequenceMatcher
from typing import Dict, List, Optional, Tuple

import openpyxl
from django.db import transaction
from django.utils import timezone

from core.models import ServiceStatus
from systems.personnel.models import PersonnelMaster
from systems.services.models import (
    ReconciliationTask,
    ServiceEventLog,
    AuditLog
)


class ReconciliationError(Exception):
    """خطأ في عملية المطابقة"""
    pass


class ReconciliationService:
    """
    محرك المطابقة - مقارنة ملف خارجي مع قاعدة البيانات
    
    التصنيفات:
    - 🟢 matched: جميع الحقول متطابقة
    - 🟡 different: اختلاف في حقل غير حساس (الحالة، المؤهل)
    - 🔴 new_in_file: موجود في الملف فقط
    - 🔵 missing_from_file: موجود في قاعدة البيانات فقط
    - ⚠️ sensitive_alert: اختلاف في الاسم/الرتبة/الرقم الوطني
    """
    
    # نسبة التشابه الدنيا للبحث التقريبي
    FUZZY_THRESHOLD = 0.85
    
    # الحقول الحساسة (لا يمكن تطبيقها مباشرة)
    SENSITIVE_FIELDS = ['full_name', 'rank', 'national_id']
    
    def __init__(self, user):
        self.user = user
    
    def create_task(self, name: str, task_type: str, 
                    source_file, key_field: str = 'military_number') -> ReconciliationTask:
        """
        إنشاء مهمة مطابقة جديدة
        
        Args:
            name: اسم المهمة
            task_type: نوع المطابقة (attendance, payroll, qualification)
            source_file: ملف Excel المصدر
            key_field: حقل الربط (military_number أو national_id)
            governorate: المحافظة (اختياري، يتم جلبه من المستخدم إذا لم يمرر)
        """
        if task_type not in ['attendance', 'payroll', 'qualification']:
            raise ReconciliationError(f'نوع مطابقة غير مدعوم: {task_type}')
        
        if key_field not in ['military_number', 'national_id']:
            raise ReconciliationError(f'حقل ربط غير مدعوم: {key_field}')
        
        
        governorate = getattr(self.user, 'profile', None) and self.user.profile.governorate
        
        task = ReconciliationTask.objects.create(
            name=name,
            governorate=governorate,
            task_type=task_type,
            source_file=source_file,
            key_field=key_field,
            status='pending',
            created_by=self.user
        )
        
        return task
    
    def process_task(self, task_id: int) -> Dict:
        """
        معالجة مهمة المطابقة
        
        Args:
            task_id: معرف المهمة
            
        Returns:
            قاموس النتائج
        """
        try:
            task = ReconciliationTask.objects.get(id=task_id)
        except ReconciliationTask.DoesNotExist:
            raise ReconciliationError(f'المهمة غير موجودة: {task_id}')
        
        if task.status == 'completed':
            raise ReconciliationError('المهمة مكتملة بالفعل')
        
        try:
            # قراءة الملف
            file_data = self._read_source_file(task.source_file, task.key_field)
            
            # جلب بيانات قاعدة البيانات الخاصة بالمحافظة فقط
            db_data = self._get_db_data(task.key_field, task.governorate_id)
            
            # المقارنة والتصنيف
            result = self._compare_data(file_data, db_data, task.key_field)
            
            # حفظ النتائج
            task.result = result
            task.status = 'completed'
            task.save(update_fields=['result', 'status'])
            
            return result
            
        except Exception as e:
            task.status = 'failed'
            task.result = {'error': str(e)}
            task.save(update_fields=['result', 'status'])
            raise ReconciliationError(f'فشل في معالجة المهمة: {str(e)}')
    
    def _read_source_file(self, source_file, key_field: str) -> List[Dict]:
        """
        قراءة ملف Excel المصدر
        
        Returns:
            قائمة قواميس لكل صف
        """
        source_file.seek(0)
        wb = openpyxl.load_workbook(source_file, data_only=True)
        ws = wb.active
        
        # قراءة العناوين
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    row_dict[header] = str(row[i]).strip() if row[i] else ''
            
            data.append(row_dict)
        
        wb.close()
        return data
    
    def _get_db_data(self, key_field: str, governorate_id: Optional[int] = None) -> Dict[str, PersonnelMaster]:
        """
        جلب بيانات الأفراد من قاعدة البيانات
        
        Returns:
            قاموس: مفتاح → PersonnelMaster
        """
        qs = PersonnelMaster.objects.select_related(
            'current_rank', 'current_status', 'central_department'
        )
        if governorate_id:
            qs = qs.filter(governorate_id=governorate_id)
            
        personnel = qs.all()
        
        db_data = {}
        for p in personnel:
            key = getattr(p, key_field, '')
            if key:
                db_data[str(key)] = p
        
        return db_data
    
    def _compare_data(self, file_data: List[Dict], 
                      db_data: Dict[str, PersonnelMaster],
                      key_field: str) -> Dict:
        """
        مقارنة بيانات الملف مع قاعدة البيانات
        
        Returns:
            نتائج المقارنة مصنفة
        """
        # خريطة أسماء الأعمدة العربية → أسماء الحقول
        field_map = {
            'الرقم العسكري': 'military_number',
            'الاسم الكامل': 'full_name',
            'الاسم': 'full_name',
            'الرتبة': 'rank',
            'الرقم الوطني': 'national_id',
            'الحالة': 'current_status',
            'الحالة الحالية': 'current_status',
            'المؤهل': 'qualification',
        }
        
        # النتائج
        matched = []
        different = []
        new_in_file = []
        sensitive_alerts = []
        
        file_keys = set()
        
        for row in file_data:
            # استخراج المفتاح
            key = row.get('الرقم العسكري', '') or row.get('الرقم الوطني', '')
            if not key:
                continue
            
            file_keys.add(key)
            
            if key not in db_data:
                # بحث تقريبي بالاسم
                name = row.get('الاسم الكامل', '') or row.get('الاسم', '')
                fuzzy_match = self._fuzzy_match_name(name, db_data) if name else None
                
                new_in_file.append({
                    'key': key,
                    'file_data': row,
                    'fuzzy_suggestion': fuzzy_match
                })
                continue
            
            # موجود في كلاهما → مقارنة
            person = db_data[key]
            differences = self._find_differences(row, person, field_map)
            
            if not differences:
                matched.append({'key': key, 'name': person.full_name})
            else:
                # فصل الاختلافات الحساسة عن غيرها
                sensitive = [d for d in differences if d['field'] in self.SENSITIVE_FIELDS]
                non_sensitive = [d for d in differences if d['field'] not in self.SENSITIVE_FIELDS]
                
                if sensitive:
                    sensitive_alerts.append({
                        'key': key,
                        'name': person.full_name,
                        'alerts': sensitive
                    })
                
                if non_sensitive:
                    different.append({
                        'key': key,
                        'name': person.full_name,
                        'personnel_id': person.pk,
                        'differences': non_sensitive
                    })
        
        # المفقودون من الملف (موجودون في DB فقط)
        missing_from_file = []
        for key, person in db_data.items():
            if key not in file_keys:
                missing_from_file.append({
                    'key': key,
                    'name': person.full_name,
                    'status': person.current_status.name if person.current_status else ''
                })
        
        return {
            'matched': matched,
            'different': different,
            'new_in_file': new_in_file,
            'missing_from_file': missing_from_file,
            'sensitive_alerts': sensitive_alerts,
            'summary': {
                'matched_count': len(matched),
                'different_count': len(different),
                'new_count': len(new_in_file),
                'missing_count': len(missing_from_file),
                'alerts_count': len(sensitive_alerts)
            }
        }
    
    def _find_differences(self, file_row: Dict, person: PersonnelMaster,
                          field_map: Dict) -> List[Dict]:
        """
        البحث عن الاختلافات بين بيانات الملف وقاعدة البيانات
        """
        differences = []
        
        # خريطة القيم من DB
        db_values = {
            'military_number': person.military_number,
            'full_name': person.full_name,
            'rank': person.current_rank.name if person.current_rank else '',
            'national_id': person.national_id or '',
            'current_status': person.current_status.name if person.current_status else '',
        }
        
        for arabic_name, field_name in field_map.items():
            if arabic_name in file_row and field_name in db_values:
                file_value = file_row[arabic_name].strip()
                db_value = str(db_values[field_name]).strip()
                
                if file_value and db_value and file_value != db_value:
                    differences.append({
                        'field': field_name,
                        'field_arabic': arabic_name,
                        'file_value': file_value,
                        'db_value': db_value
                    })
        
        return differences
    
    def _fuzzy_match_name(self, name: str, 
                          db_data: Dict[str, PersonnelMaster]) -> Optional[Dict]:
        """
        بحث تقريبي بالاسم (Levenshtein distance)
        
        Returns:
            أقرب تطابق (None إذا لم يوجد)
        """
        best_match = None
        best_ratio = 0.0
        
        for key, person in db_data.items():
            ratio = SequenceMatcher(None, name, person.full_name).ratio()
            if ratio > best_ratio and ratio >= self.FUZZY_THRESHOLD:
                best_ratio = ratio
                best_match = {
                    'key': key,
                    'name': person.full_name,
                    'similarity': round(ratio * 100, 1)
                }
        
        return best_match
    
    @transaction.atomic
    def apply_change(self, task_id: int, record_index: int,
                     source: str = 'file') -> Dict:
        """
        تطبيق تغيير من نتائج المطابقة
        
        Args:
            task_id: معرف المهمة
            record_index: فهرس السجل في قائمة different
            source: مصدر القيمة الصحيحة ('file' أو 'db')
            
        Returns:
            نتيجة التطبيق
        """
        try:
            task = ReconciliationTask.objects.get(id=task_id, status='completed')
        except ReconciliationTask.DoesNotExist:
            raise ReconciliationError('المهمة غير موجودة أو لم تكتمل')
        
        result = task.result
        if not result or 'different' not in result:
            raise ReconciliationError('لا توجد نتائج مطابقة')
        
        different = result['different']
        if record_index >= len(different):
            raise ReconciliationError(f'فهرس غير صالح: {record_index}')
        
        record = different[record_index]
        personnel_id = record.get('personnel_id')
        
        if not personnel_id:
            raise ReconciliationError('لا يمكن تطبيق التغيير: معرف الفرد غير متوفر')
        
        try:
            personnel = PersonnelMaster.objects.select_related(
                'current_status'
            ).get(pk=personnel_id)
        except PersonnelMaster.DoesNotExist:
            raise ReconciliationError('الفرد غير موجود')
        
        applied_changes = []
        
        for diff in record['differences']:
            field = diff['field']
            
            # لا نطبق التغييرات الحساسة
            if field in self.SENSITIVE_FIELDS:
                continue
            
            if field == 'current_status':
                new_value = diff['file_value'] if source == 'file' else diff['db_value']
                old_value = personnel.current_status.name if personnel.current_status else ''
                
                if new_value != old_value:
                    try:
                        new_status = ServiceStatus.objects.get(name=new_value)
                    except ServiceStatus.DoesNotExist:
                        continue
                    
                    # التحقق من قاعدة الانتقال
                    try:
                        from systems.services.application.services.atomic_service import AtomicService
                        AtomicService.validate_transition(
                            personnel.current_status, new_status
                        )
                    except Exception as e:
                        # تسجيل التحذير وتجاوز هذا التغيير
                        continue
                    
                    personnel.current_status = new_status
                    personnel.save(update_fields=['current_status'])
                    
                    ServiceEventLog.objects.create(
                        personnel=personnel,
                        security_admin=personnel.security_admin,
                        event_date=date.today(),
                        service_month=date.today().strftime('%Y-%m'),
                        field_name='current_status',
                        old_value=old_value,
                        new_value=new_value,
                        created_by=self.user
                    )
                    
                    applied_changes.append({
                        'field': field,
                        'old_value': old_value,
                        'new_value': new_value
                    })
        
        # AuditLog
        if applied_changes:
            AuditLog.objects.create(
                user=self.user,
                action='UPDATE',
                model_name='PersonnelMaster',
                object_id=str(personnel.pk),
                old_data={'source': 'reconciliation', 'task_id': task_id},
                new_data={'changes': applied_changes}
            )
        
        return {
            'success': True,
            'military_number': personnel.military_number,
            'applied_changes': applied_changes
        }
