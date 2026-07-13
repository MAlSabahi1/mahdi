"""
Report Generation Service - خدمة توليد التقارير
المرحلة 4 - المهمة 4.5: توليد التقارير الرسمية
"""
import os
import uuid
from io import BytesIO
from datetime import date

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.conf import settings
from django.db.models import Count, Q

from systems.personnel.models import PersonnelMaster
from systems.services.models import (
    AuditLog, DirectorateCompliance,
    ReportTemplate, FormEventLog, StatusChangeForm,
)
from core.models import CentralDepartment


class ReportGenerationService:
    """خدمة توليد التقارير"""
    
    def __init__(self, user):
        self.user = user
    
    def generate(self, template_slug, filters, output_format='excel'):
        """
        توليد تقرير حسب القالب

        Args:
            template_slug: رمز القالب
            filters: معايير التصفية (month, directorate_id, etc.)
            output_format: تنسيق الإخراج ('excel' أو 'pdf')
        """
        generators = {
            'personnel_summary': self._generate_personnel_summary,
            'department_strength': self._generate_directorate_strength,
            'monthly_changes': self._generate_monthly_changes,
            'rejections_report': self._generate_rejections_report,
        }
        
        generator = generators.get(template_slug)
        if not generator:
            raise ValueError(f'قالب غير مدعوم: {template_slug}')
        
        content = generator(filters)
        
        # حفظ الملف
        file_name = f'{template_slug}_{uuid.uuid4().hex[:8]}.xlsx'
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports', file_name)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'wb') as f:
            f.write(content)
        
        # تسجيل في AuditLog
        AuditLog.objects.create(
            user=self.user,
            action='EXPORT',
            model_name='ReportTemplate',
            object_id=template_slug,
            new_data={'filters': filters, 'format': output_format},
        )
        
        return {
            'file_path': file_path,
            'file_name': file_name,
            'file_url': f'{settings.MEDIA_URL}reports/{file_name}',
        }
    
    def _generate_personnel_summary(self, filters):
        """نموذج 1: ملخص الأفراد"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ملخص الأفراد'
        ws.sheet_view.rightToLeft = True
        
        # عناوين
        headers = ['الرقم العسكري', 'الاسم الكامل', 'الرتبة', 'الإدارة', 'الحالة', 'المؤهل']
        header_font = Font(bold=True, size=12)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        qs = PersonnelMaster.objects.select_related(
            'current_rank', 'central_department', 'current_status', 'qualification'
        )
        
        dept_id = filters.get('directorate_id')
        if dept_id:
            qs = qs.filter(central_department_id=dept_id)
        
        for row_num, p in enumerate(qs, 2):
            ws.cell(row=row_num, column=1, value=p.military_number)
            ws.cell(row=row_num, column=2, value=p.full_name)
            ws.cell(row=row_num, column=3, value=p.current_rank.name if p.current_rank else '')
            ws.cell(row=row_num, column=4, value=p.central_department.name if p.central_department else '')
            ws.cell(row=row_num, column=5, value=p.current_status.name if p.current_status else '')
            ws.cell(row=row_num, column=6, value=p.qualification.name if p.qualification else '')
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _generate_directorate_strength(self, filters):
        """نموذج 2: قوة المديريات"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'قوة المديريات'
        ws.sheet_view.rightToLeft = True
        
        headers = ['المديرية', 'العدد الكلي', 'على رأس العمل', 'إجازة', 'منقطع', 'أخرى']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True, size=12)
        
        directorates = CentralDepartment.objects.filter(is_active=True)
        
        for row_num, dept in enumerate(directorates, 2):
            personnel = PersonnelMaster.objects.filter(central_department=dept)
            total = personnel.count()
            active = personnel.filter(
                current_status__classification='active_full'
            ).count()
            leave = personnel.filter(
                current_status__classification='inactive_temp'
            ).count()
            # الغائبون: حالة 'absent' فقط (منفصلة عن inactive_temp)
            absent = personnel.filter(
                current_status__classification='absent'
            ).count()
            other = max(total - active - leave - absent, 0)
            
            ws.cell(row=row_num, column=1, value=dept.name)
            ws.cell(row=row_num, column=2, value=total)
            ws.cell(row=row_num, column=3, value=active)
            ws.cell(row=row_num, column=4, value=leave)
            ws.cell(row=row_num, column=5, value=absent)
            ws.cell(row=row_num, column=6, value=other)
        
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _generate_monthly_changes(self, filters):
        """نموذج 3: التغييرات الشهرية — من استمارات الخدمات المعتمدة"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'التغييرات الشهرية'
        ws.sheet_view.rightToLeft = True

        headers = ['م', 'الرقم العسكري', 'الاسم', 'الرتبة', 'الإدارة',
                   'نوع الخدمة', 'الحالة السابقة', 'الحالة الجديدة', 'تاريخ الاعتماد']
        header_font = Font(bold=True, size=11)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # جلب الاستمارات المعتمدة خلال الشهر المحدد
        month = filters.get('month', date.today().strftime('%Y-%m'))
        year, mon = month.split('-') if '-' in month else (date.today().year, date.today().month)

        qs = StatusChangeForm.objects.select_related(
            'personnel', 'personnel__current_rank',
            'personnel__central_department',
            'from_status', 'to_status', 'service_catalog',
        ).filter(
            status='approved',
            updated_at__year=int(year),
            updated_at__month=int(mon),
        )

        dept_id = filters.get('directorate_id')
        if dept_id:
            qs = qs.filter(personnel__central_department_id=dept_id)

        for row_num, form in enumerate(qs, 2):
            p = form.personnel
            service_name = (
                form.service_catalog.name_ar if form.service_catalog
                else form.get_form_type_display() if hasattr(form, 'get_form_type_display')
                else form.form_type
            )
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=p.military_number if p else '')
            ws.cell(row=row_num, column=3, value=p.full_name if p else '')
            ws.cell(row=row_num, column=4, value=p.current_rank.name if p and p.current_rank else '')
            ws.cell(row=row_num, column=5, value=p.central_department.name if p and p.central_department else '')
            ws.cell(row=row_num, column=6, value=service_name)
            ws.cell(row=row_num, column=7, value=form.from_status.name if form.from_status else '')
            ws.cell(row=row_num, column=8, value=form.to_status.name if form.to_status else '')
            ws.cell(row=row_num, column=9, value=str(form.updated_at.date()) if form.updated_at else '')

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def _generate_rejections_report(self, filters):
        """نموذج 4: تقرير المرفوضات — من FormEventLog (استمارات الخدمات)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'المرفوضات'
        ws.sheet_view.rightToLeft = True

        headers = ['م', 'الرقم العسكري', 'الاسم', 'الرتبة', 'الإدارة',
                   'نوع الخدمة', 'الحالة المطلوبة', 'سبب الرفض', 'رفض بواسطة', 'التاريخ']
        header_font = Font(bold=True, size=11)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # جلب الاستمارات المرفوضة من FormEventLog
        qs = FormEventLog.objects.select_related(
            'form', 'form__personnel', 'form__personnel__current_rank',
            'form__personnel__central_department',
            'form__to_status', 'form__service_catalog',
            'performed_by',
        ).filter(action='rejected')

        month = filters.get('month')
        if month and '-' in month:
            year, mon = month.split('-')
            qs = qs.filter(
                created_at__year=int(year),
                created_at__month=int(mon),
            )

        dept_id = filters.get('directorate_id')
        if dept_id:
            qs = qs.filter(form__personnel__central_department_id=dept_id)

        for row_num, event in enumerate(qs, 2):
            form = event.form
            p = form.personnel if form else None
            service_name = (
                form.service_catalog.name_ar if form and form.service_catalog
                else form.get_form_type_display() if form and hasattr(form, 'get_form_type_display')
                else (form.form_type if form else '')
            )
            to_status_name = form.to_status.name if form and form.to_status else ''
            rejection_reason = event.notes.replace('سبب الرفض: ', '') if event.notes else ''

            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=p.military_number if p else '')
            ws.cell(row=row_num, column=3, value=p.full_name if p else '')
            ws.cell(row=row_num, column=4, value=p.current_rank.name if p and p.current_rank else '')
            ws.cell(row=row_num, column=5, value=p.central_department.name if p and p.central_department else '')
            ws.cell(row=row_num, column=6, value=service_name)
            ws.cell(row=row_num, column=7, value=to_status_name)
            ws.cell(row=row_num, column=8, value=rejection_reason)
            ws.cell(row=row_num, column=9, value=event.performed_by.username if event.performed_by else '')
            ws.cell(row=row_num, column=10, value=str(event.created_at.date()) if event.created_at else '')

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
