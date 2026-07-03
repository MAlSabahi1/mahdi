"""
Strict Initial Import Service — خدمة الاستيراد التأسيسي الصارم
═════════════════════════════════════════════════════════════
الهدف: استيراد البيانات من Excel أثناء تأسيس النظام فقط بصرامة تامة (100% مطابقة).
لا يسمح بتعيينات عشوائية ولا يخمن الرتب أو الأقسام.

الخطوات:
1. Validate: قراءة الملف وإرجاع تقرير بالأخطاء (بدون حفظ أي شيء في Master).
2. Commit: حفظ البيانات النهائية إذا اجتازت الفحص بالكامل 100%.
"""
import uuid
import re
from datetime import datetime
from typing import Dict, List, Any, Tuple

import openpyxl
from django.db import transaction
from django.utils import timezone
from django.core.exceptions import ValidationError

from core.models import (
    Rank, ServiceStatus, SecurityAdministration,
    CentralDepartment, Branch, DistrictPolice,
    Division, Unit, ForceType, Qualification,
    JobCategory, JobTitle, Position
)
from systems.personnel.models import (
    PersonnelMaster, RawDataImport, SuggestedCorrection,
    HistoricalMonthlyVariables
)


def format_validation_error(e: ValidationError) -> str:
    msg = e.message
    if hasattr(e, 'params') and e.params:
        try:
            msg = msg % e.params
        except Exception:
            pass
    return str(msg)


# ── جداول تصحيح الرتب (للمطابقة الذكية قبل الرفض) ──
RANK_MAP = {
    'لواء': 'لواء', 'عميد': 'عميد', 'عقيد': 'عقيد',
    'مقدم': 'مقدم', 'رائد': 'رائد', 'نقيب': 'نقيب',
    'ملازم أول': 'ملازم أول', 'ملازم ثاني': 'ملازم ثاني',
    'مساعد1': 'مساعد 1', 'مساعد 1': 'مساعد 1', 'مساعد': 'مساعد 1',
    'مساعد2': 'مساعد 2', 'مساعد 2': 'مساعد 2',
    'رقيب1': 'رقيب 1', 'رقيب 1': 'رقيب 1',
    'رقيب2': 'رقيب 2', 'رقيب 2': 'رقيب 2',
    'عريف': 'عريف', 'جندي': 'جندي', 'مجند': 'جندي',
    'حارس': 'حارس', 'مدني': 'مدني', 'متعاقد': 'متعاقد',
}

# ── جداول تصحيح الحالات — مطابقة للقيم الفعلية في قاعدة البيانات ──
STATUS_MAP = {
    # الحالات بالصيغة المُعرَّفة (بـ ال)
    'الجرحى': 'الجرحى',
    'الشهداء': 'الشهداء',
    'الوفيات': 'الوفيات',
    'الأسرى': 'الأسرى',
    'السجناء': 'السجناء',
    'الإجازات': 'الإجازات',
    'الغياب المستمر': 'الغياب المستمر',
    'المتقاعدين': 'المتقاعدين',
    'المصدرين إلى الوزارة': 'المصدرين إلى الوزارة',
    'المفرغين في الجبهات': 'المفرغين في الجبهات',
    'المفرغين للدراسة': 'المفرغين للدراسة',
    'المفرغين للمرافقة': 'المفرغين للمرافقة',
    'المفقودين': 'المفقودين',
    'المفقودين (بصك شرعي)': 'المفقودين (بصك شرعي)',
    'الملتحقين بالعدوان': 'الملتحقين بالعدوان',
    'المنتدبين لدى جهات': 'المنتدبين لدى جهات',
    'المنقطعين - فرار': 'المنقطعين - فرار',
    'الوفيات': 'الوفيات',
    'الأمراض والمصابين': 'الأمراض والمصابين',
    'إنهاء المدة القانونية': 'إنهاء المدة القانونية',
    'بلوغ السن القانوني': 'بلوغ السن القانوني',
    'تعمل في الميدان': 'تعمل في الميدان',
    'عدم لياقة': 'عدم لياقة',
    # مرادفات شائعة → القيمة الصحيحة في DB
    'عاملين': 'تعمل في الميدان',
    'عامل': 'تعمل في الميدان',
    'قوة عاملة فعلية': 'تعمل في الميدان',
    'قوة عاملة/فعلية': 'تعمل في الميدان',
    'بدون عمل': 'بدون عمل',
    'قوة احتياط': 'قوة احتياطية',
    'قوة احتياطية': 'قوة احتياطية',
    'دارسين': 'المفرغين للدراسة',
    'دراسة': 'المفرغين للدراسة',
    'منتدب': 'المنتدبين لدى جهات',
    'منتدبين': 'المنتدبين لدى جهات',
    'مصدر': 'المصدرين إلى الوزارة',
    'مصدرين': 'المصدرين إلى الوزارة',
    'منزل': 'المصدرين إلى الوزارة',
    'معارين': 'المنتدبين لدى جهات',
    'معيات': 'المنتدبين لدى جهات',
    'غياب': 'الغياب المستمر',
    'غياب مستمر': 'الغياب المستمر',
    'جريح': 'الجرحى',
    'جرحى': 'الجرحى',
    'أسير': 'الأسرى',
    'أسرى': 'الأسرى',
    'سجين': 'السجناء',
    'سجناء': 'السجناء',
    'إجازة': 'الإجازات',
    'إجازات': 'الإجازات',
    'شهيد': 'الشهداء',
    'شهداء': 'الشهداء',
    'متوفى': 'الوفيات',
    'المعيات': 'المعيات',
    'معيات': 'المعيات',
    'وفاة': 'الوفيات',
    'وفيات': 'الوفيات',
    'متقاعد': 'المتقاعدين',
    'تقاعد': 'المتقاعدين',
    'متقاعدين': 'المتقاعدين',
    'مفصول': 'المنقطعين - فرار',
    'مفصولين': 'المنقطعين - فرار',
    'فرار': 'المنقطعين - فرار',
    'منقطع': 'المنقطعين - فرار',
    'منقطعين - فرار': 'المنقطعين - فرار',
    'ملتحق بالعدوان': 'الملتحقين بالعدوان',
    'ملتحقين بالعدوان': 'الملتحقين بالعدوان',
    'مهام أمنية': 'تعمل في الميدان',
    'خدمة أمنية': 'تعمل في الميدان',
    'مفرغ للدراسة': 'المفرغين للدراسة',
    'مفرغ في الجبهة': 'المفرغين في الجبهات',
    'مفرغ للمرافقة': 'المفرغين للمرافقة',
    'مفقود': 'المفقودين',
    'مريض': 'الأمراض والمصابين',
    'مصاب': 'الأمراض والمصابين',
}

# ── تصنيف القوة ──
FORCE_TYPE_MAP = {
    'القوة الأساسي- أفراد': 'القوة الأساسي - أفراد',
    'القوة الأساسي - أفراد': 'القوة الأساسي - أفراد',
    'لجان أمنية- أفراد': 'لجان أمنية - أفراد',
    'لجان أمنية - أفراد': 'لجان أمنية - أفراد',
    'قوة- مستجدين جديد': 'قوة - مستجدين جديد',
    'قوة إداريا فقط': 'إداريا فقط',
    'إداريا فقط': 'إداريا فقط',
    'قوة منزلة- خارج الصرف': 'قوة منزلة - خارج الصرف',
    'عاملين': 'القوة الأساسي - أفراد',
}

# ── الإدارات الأمنية — مرادفات → الاسم الرسمي في قاعدة البيانات ──
UNIT_MAP = {
    # المحافظة الرئيسية
    'المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'شرطة المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'شرطة المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'شرطة محافظة المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'امن المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'أمن المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'امن محافظة المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'إدارة أمن المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'ادارة امن المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    'إدارة أمن المحافظة الرئيسية': 'إدارة أمن محافظة المحافظة الرئيسية',
    # المحافظة المركزية
    'المحافظة المركزية': 'إدارة أمن المحافظة المركزية',
    'امن المحافظة المركزية': 'إدارة أمن المحافظة المركزية',
    'شرطة المحافظة المركزية': 'إدارة أمن المحافظة المركزية',
    # عدن
    'عدن': 'إدارة أمن عدن',
    'امن عدن': 'إدارة أمن عدن',
    # تعز
    'تعز': 'إدارة أمن تعز',
    'امن تعز': 'إدارة أمن تعز',
    # أمانة العاصمة
    'أمانة العاصمة': 'إدارة أمن أمانة العاصمة',
    'امانة': 'إدارة أمن أمانة العاصمة',
    'امانة العاصمة': 'إدارة أمن أمانة العاصمة',
    # الحديدة
    'الحديدة': 'إدارة أمن الحديدة',
    'حديدة': 'إدارة أمن الحديدة',
    # حضرموت
    'حضرموت': 'إدارة أمن حضرموت',
    # إب
    'إب': 'إدارة أمن إب',
    'اب': 'إدارة أمن إب',
    # ذمار
    'ذمار': 'إدارة أمن ذمار',
    # حجة
    'حجة': 'إدارة أمن حجة',
    # شبوة
    'شبوة': 'إدارة أمن شبوة',
    # أبين
    'أبين': 'إدارة أمن أبين',
    'ابين': 'إدارة أمن أبين',
    # البيضاء
    'البيضاء': 'إدارة أمن البيضاء',
    # الجوف
    'الجوف': 'إدارة أمن الجوف',
    # الضالع
    'الضالع': 'إدارة أمن الضالع',
    # المحويت
    'المحويت': 'إدارة أمن المحويت',
    # المهرة
    'المهرة': 'إدارة أمن المهرة',
    # ريمة
    'ريمة': 'إدارة أمن ريمة',
    # سقطرى
    'سقطرى': 'إدارة أمن سقطرى',
    'سقطرا': 'إدارة أمن سقطرى',
    # صعدة
    'صعدة': 'إدارة أمن صعدة',
    # عمران
    'عمران': 'إدارة أمن عمران',
    # لحج
    'لحج': 'إدارة أمن لحج',
    # المحافظة الرئيسية
    'المحافظة الرئيسية': 'إدارة أمن المحافظة الرئيسية',
    'المحافظة الرئيسية': 'إدارة أمن المحافظة الرئيسية',
    'شرطة المحافظة الرئيسية': 'إدارة أمن المحافظة الرئيسية',
    'شرطة محافظة المحافظة الرئيسية': 'إدارة أمن المحافظة الرئيسية',
}

# ── حالة النفقات ──
EXPENSE_MAP = {
    'لديه نفقات': 'has_expenses',
    'نفقات': 'expenses',
    'بدون نفقات': 'no_expenses',
}

class StrictImportError(Exception):
    pass

class StrictInitialImportService:
    
    EXPECTED_COLUMNS = [
        'الرقم العسكري',
        'الرقم العسكري القديم',
        'الاسم',
        'الاسم الكامل',
        'تصحيح الاسم من واقع البطاقة',
        'الرتبة',
        'الرقم الوطني',
        'الحالة',
        'نوع الحالة',
        'نوع العمل',
        'الفئة',
        'المنصب',
        'تصنيف القوة',
        'المؤهل',
        'الوحدة',
        'الإدارة_السرية',
        'القسم_فرع السرية',
        'تاريخ الميلاد',
        'تاريخ الألتحاق',
        'تاريخ صدور القرار',
        'تاريخ التصدور الينا',
        'رقم التليفون',
        'حالة النفقات',
        'التعيينات',
        'الملاحظات'
    ]

    def __init__(self):
        self._load_dictionaries()

    def _load_dictionaries(self):
        """تحميل كل القواميس في الذاكرة لتسريع الفحص الصارم"""
        from collections import defaultdict
        
        self.ranks = {r.name.strip(): r for r in Rank.objects.all()}
        self.statuses = {s.name.strip(): s for s in ServiceStatus.objects.all()}
        self.security_admins = {s.name.strip(): s for s in SecurityAdministration.objects.all()}
        
        self.central_depts = defaultdict(list)
        for c in CentralDepartment.objects.all():
            self.central_depts[c.name.strip()].append(c)
            
        self.branches = defaultdict(list)
        for b in Branch.objects.all():
            self.branches[b.name.strip()].append(b)
            
        self.district_police = defaultdict(list)
        for d in DistrictPolice.objects.all():
            self.district_police[d.name.strip()].append(d)
            
        self.divisions = defaultdict(list)
        for d in Division.objects.all():
            self.divisions[d.name.strip()].append(d)
        
        self.force_types = {f.name.strip(): f for f in ForceType.objects.all()}
        self.qualifications = {q.name.strip(): q for q in Qualification.objects.all()}
        self.job_categories = {j.name.strip(): j for j in JobCategory.objects.all()}
        self.job_titles = {j.name.strip(): j for j in JobTitle.objects.all()}
        self.positions = {p.name.strip(): p for p in Position.objects.all()}

    def _clean(self, val) -> str:
        if val is None:
            return ""
        val_str = str(val).strip()
        # تحويل الأرقام المشرقية (العربية) إلى أرقام إنجليزية لضمان صحة البيانات في كل الأعمدة (هواتف، أرقام عسكرية...)
        arabic_digits = "٠١٢٣٤٥٦٧٨٩"
        english_digits = "0123456789"
        return val_str.translate(str.maketrans(arabic_digits, english_digits))

    def _parse_date(self, val):
        if not val:
            return None
        if isinstance(val, datetime):
            return val.date()
        val = str(val).strip()
        
        # تجاهل الوقت إذا كان موجوداً (مثال: 1977-07-07 00:00:00)
        if len(val) > 10 and val[10] in (' ', 'T'):
            val = val[:10]
            
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%m/%d/%Y'):
            try:
                return datetime.strptime(val, fmt).date()
            except (ValueError, TypeError):
                continue
        return None

    def extract_preview(self, file_content: bytes) -> Dict[str, Any]:
        """
        قراءة سريعة لأسماء الأعمدة وأول 10 صفوف من الملف لعرضها كمعاينة (Preview) للربط اليدوي.
        """
        import io
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            raise StrictImportError(f"فشل قراءة ملف الإكسل: {str(e)}")

        rows = list(ws.iter_rows(values_only=True, max_row=11))
        if len(rows) < 2:
            raise StrictImportError("الملف فارغ أو لا يحتوي على بيانات كافية للمعاينة")

        headers = [self._clean(h) for h in rows[0]]
        preview_rows = []
        for row in rows[1:]:
            preview_rows.append([self._clean(cell) for cell in row])

        return {
            "headers": headers,
            "rows": preview_rows
        }

    def parse_file_to_dicts(self, file_content: bytes, column_mapping: Dict[str, str] = None) -> List[Dict[str, Any]]:
        import io
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            raise StrictImportError(f"فشل قراءة ملف الإكسل: {str(e)}")

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise StrictImportError("الملف فارغ أو لا يحتوي على بيانات")

        headers = [self._clean(h) for h in rows[0]]
        
        # تطبيق الربط اليدوي (Column Mapping) إذا توفر
        ignored_indices = set()
        if column_mapping:
            for i, h in enumerate(headers):
                if h in column_mapping:
                    if column_mapping[h] == '__ignore__':
                        ignored_indices.add(i)
                    elif column_mapping[h]:
                        headers[i] = column_mapping[h]
        
        def _normalize_arabic(text):
            if not text:
                return "", ""
            t = str(text)
            t = re.sub(r'[أإآ]', 'ا', t)
            t = t.replace('ة', 'ه').replace('ي', 'ى')
            # Remove generic words for unit matching
            t_unit = re.sub(r'(إداره|اداره|أمن|امن|شرطه|محافظه|امانه|العاصمه)', '', t).strip()
            return t.strip(), t_unit
            
        # التحقق من وجود الأعمدة المطلوبة كحد أدنى
        missing_cols = []
        if 'الرقم العسكري' not in headers and 'الرقم العسكري القديم' not in headers:
            missing_cols.append('الرقم العسكري')
        if 'الاسم' not in headers and 'الاسم الكامل' not in headers and 'الأسم' not in headers:
            missing_cols.append('الاسم')
        
        if missing_cols:
            raise StrictImportError(f"الأعمدة التالية مفقودة أساسياً: {', '.join(missing_cols)}")
            
        data_dicts = []
        for row in rows[1:]:
            if not row or not any(row):
                continue
            
            row_dict = {}
            for i, val in enumerate(row):
                if i in ignored_indices:
                    continue
                h = headers[i]
                
                val_str = self._clean(val)
                val_norm, val_unit_norm = _normalize_arabic(val_str)
                
                if h == 'الوحدة':
                    # Dynamic match against DB
                    if val_str not in self.security_admins:
                        for db_name in self.security_admins:
                            _, db_unit_norm = _normalize_arabic(db_name)
                            if val_unit_norm and db_unit_norm and (val_unit_norm in db_unit_norm or db_unit_norm in val_unit_norm):
                                val = db_name
                                break
                elif h == 'الحالة':
                    if val_str in STATUS_MAP:
                        val = STATUS_MAP[val_str]
                elif h == 'الرتبة':
                    mapped = RANK_MAP.get(val_str, val_str)
                    if mapped in self.ranks:
                        val = mapped
                    elif val_str not in self.ranks:
                        # Try normalized match
                        for db_name in self.ranks:
                            db_norm, _ = _normalize_arabic(db_name)
                            if db_norm == val_norm:
                                val = db_name
                                break
                elif h in ('جهة_العمل', 'الإدارة_السرية'):
                    # Match known generic prefixes
                    if val_str.startswith('مديرية '):
                        val_str = val_str.replace('مديرية ', 'شرطة مديرية ', 1)
                        val_norm, _ = _normalize_arabic(val_str)
                        
                    # Try to find a dynamic match in central depts, branches, or district police using normalized names
                    found = False
                    for db_dict in (self.central_depts, self.branches, self.district_police):
                        if val_str in db_dict:
                            val = val_str
                            found = True
                            break
                        
                        # Try normalized match
                        for db_name in db_dict:
                            db_norm, _ = _normalize_arabic(db_name)
                            if db_norm == val_norm or db_norm.endswith(val_norm) or val_norm.endswith(db_norm):
                                val = db_name
                                found = True
                                break
                        if found:
                            break
                elif h in ('الفئة', 'نوع العمل', 'المنصب', 'المؤهل', 'تصنيف القوة'):
                    db_dict = None
                    if h == 'الفئة':
                        db_dict = self.job_categories
                    elif h == 'نوع العمل':
                        db_dict = self.job_titles
                    elif h == 'المنصب':
                        db_dict = self.positions
                    elif h == 'المؤهل':
                        db_dict = self.qualifications
                    elif h == 'تصنيف القوة':
                        db_dict = self.force_types
                        mapped = FORCE_TYPE_MAP.get(val_str, val_str)
                        if mapped in db_dict:
                            val = mapped
                            db_dict = None  # skip fuzzy since we matched
                        
                    if db_dict and val_str not in db_dict:
                        for db_name in db_dict:
                            db_norm, _ = _normalize_arabic(db_name)
                            if db_norm == val_norm:
                                val = db_name
                                break
                elif h == 'القسم_فرع السرية' or h == 'القسم':
                    if val_str and val_str not in self.divisions:
                        for db_name in self.divisions:
                            db_norm, _ = _normalize_arabic(db_name)
                            if db_norm == val_norm:
                                val = db_name
                                break
                
                row_dict[h] = val
                
            data_dicts.append(row_dict)
            
        return data_dicts

    def validate_file(self, file_content: bytes, column_mapping: Dict[str, str] = None) -> Dict[str, Any]:
        """
        الفحص الصارم من ملف الإكسل مباشرة (يقرأ الملف ثم يحيله إلى validate_data).
        """
        rows_data = self.parse_file_to_dicts(file_content, column_mapping)
        return self.validate_data(rows_data)

    def validate_data(self, rows_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        الفحص الصارم لقائمة من البيانات (List of dicts). 
        يعيد تقرير الأخطاء بالكامل ومعه بيانات الصف لتعديلها في الواجهة.
        """
        if not rows_data:
            raise StrictImportError("لا توجد بيانات للفحص")

        errors = []
        valid_rows = 0
        
        headers = list(rows_data[0].keys())
        
        # كشف أسماء الأعمدة المتوفرة لتفادي أخطاء المفاتيح
        col_mil = 'الرقم العسكري' if 'الرقم العسكري' in headers else ('الرقم العسكري القديم' if 'الرقم العسكري القديم' in headers else '')
        col_name = 'الاسم' if 'الاسم' in headers else ('الاسم الكامل' if 'الاسم الكامل' in headers else 'الأسم')
        col_rank = 'الرتبة' if 'الرتبة' in headers else ''
        col_status = 'الحالة' if 'الحالة' in headers else ''
        col_status_type = 'نوع الحالة' if 'نوع الحالة' in headers else ''
        col_national = 'الرقم الوطني' if 'الرقم الوطني' in headers else ''
        col_unit = 'الوحدة' if 'الوحدة' in headers else ''
        col_dept = 'الإدارة_السرية' if 'الإدارة_السرية' in headers else ''
        col_div = 'القسم_فرع السرية' if 'القسم_فرع السرية' in headers else ''

        # فحص أرقام عسكرية مكررة في الملف نفسه
        file_military_numbers = {}

        for idx, row_dict in enumerate(rows_data, start=2):
            row_errors = []

            # 1. الرقم العسكري بأنواعه
            mil_correct = self._clean(row_dict.get('الرقم العسكري الصحيح', ''))
            mil_normal = self._clean(row_dict.get('الرقم العسكري', ''))
            mil_old = self._clean(row_dict.get('الرقم العسكري القديم', ''))
            
            # إذا كان هناك رقم صحيح ومختلف عن الرقم العادي، نعتبره خطأ يحتاج مراجعة
            if mil_correct and mil_normal and mil_correct != mil_normal:
                row_errors.append({
                    "field": "الرقم العسكري الصحيح", 
                    "message": f"يوجد اختلاف بين الرقم العسكري ({mil_normal}) والرقم الصحيح ({mil_correct}). يرجى المراجعة والدمج."
                })
            
            # نحتفظ بالرقم الأساسي للتحقق من وجود رقم على الأقل (نستخدم العادي، ثم الصحيح، ثم القديم)
            mil_num = mil_normal or mil_correct or mil_old
            
            if not mil_num:
                row_errors.append({"field": "الرقم العسكري", "message": "الرقم العسكري (أو القديم/الصحيح) مطلوب"})

            # فحص كل حقل موجود على حدة لضمان صحة الطول ومنع التكرار الشامل في الملف
            for col_key, col_val in [('الرقم العسكري الصحيح', mil_correct), ('الرقم العسكري', mil_normal), ('الرقم العسكري القديم', mil_old)]:
                if col_val:
                    if not re.match(r'^\d{7}$', col_val):
                        row_errors.append({"field": col_key, "message": f"{col_key} يجب أن يكون 7 أرقام بالضبط (المُدخَل: {col_val})"})
                    elif col_val in file_military_numbers:
                        dup_idx, dup_col = file_military_numbers[col_val]
                        # لا نعتبره تكراراً إذا كان في نفس الصف (مثلاً الرقم العسكري يساوي الرقم الصحيح)
                        if dup_idx != idx:
                            row_errors.append({"field": col_key, "message": f"{col_key} ({col_val}) مكرر مع '{dup_col}' في الصف رقم ({dup_idx}) في هذا الملف"})
                    else:
                        file_military_numbers[col_val] = (idx, col_key)

            # 2. الاسم وتصحيح الاسم
            full_name = self._clean(row_dict.get(col_name, ''))
            if full_name:
                # تصحيح صامت للأسماء الموصولة (عز_الدين -> عز الدين)
                full_name = full_name.replace('_', ' ').replace('-', ' ')
                full_name = re.sub(r'\s+', ' ', full_name).strip()
                row_dict[col_name] = full_name
                
                # التحقق الصارم للاسم الأساسي
                if not re.match(r'^[\u0621-\u064A\s]+$', full_name):
                    row_errors.append({"field": col_name, "message": "الاسم يجب أن يحتوي على حروف عربية فقط (يمنع استخدام الأرقام أو الرموز)"})
                elif len(full_name.split()) < 2:
                    row_errors.append({"field": col_name, "message": "الاسم يجب أن يكون ثنائياً على الأقل"})

            if not full_name:
                row_errors.append({"field": col_name, "message": "الاسم مطلوب"})

            corrected_name = self._clean(row_dict.get('تصحيح الاسم من واقع البطاقة', ''))
            if corrected_name:
                # تصحيح صامت
                corrected_name = corrected_name.replace('_', ' ').replace('-', ' ')
                corrected_name = re.sub(r'\s+', ' ', corrected_name).strip()
                row_dict['تصحيح الاسم من واقع البطاقة'] = corrected_name
                
                # التحقق من أن الاسم يحتوي على حروف عربية ومسافات فقط (لا أرقام ولا رموز كـ + وغيرها)
                if not re.match(r'^[\u0621-\u064A\s]+$', corrected_name):
                    row_errors.append({"field": "تصحيح الاسم من واقع البطاقة", "message": "الاسم المصحح يجب أن يحتوي على حروف عربية فقط (يمنع استخدام الأرقام أو الرموز)"})
                elif len(corrected_name.split()) < 2:
                    row_errors.append({"field": "تصحيح الاسم من واقع البطاقة", "message": "الاسم المصحح يجب أن يكون اسماً ثنائياً على الأقل"})

            # 3. الرقم الوطني، الهاتف، وتواريخ الميلاد والالتحاق
            nat_id = self._clean(row_dict.get(col_national, ''))
            if nat_id:
                from core.validators import validate_national_id
                try:
                    validate_national_id(nat_id)
                except ValidationError as e:
                    row_errors.append({"field": col_national, "message": format_validation_error(e)})

            col_phone = 'رقم التليفون' if 'رقم التليفون' in headers else ''
            phone = self._clean(row_dict.get(col_phone, ''))
            if phone:
                from core.validators import validate_phone_number
                try:
                    validate_phone_number(phone)
                except ValidationError as e:
                    row_errors.append({"field": col_phone, "message": format_validation_error(e)})

            col_birth = 'تاريخ الميلاد' if 'تاريخ الميلاد' in headers else ''
            col_join = 'تاريخ الألتحاق' if 'تاريخ الألتحاق' in headers else ('تاريخ الالتحاق' if 'تاريخ الالتحاق' in headers else '')
            
            birth_date_obj = None
            if col_birth:
                birth_str = self._clean(row_dict.get(col_birth, ''))
                if birth_str:
                    birth_date_obj = self._parse_date(birth_str)
                    if not birth_date_obj:
                        row_errors.append({"field": col_birth, "message": "تاريخ الميلاد غير صالح"})
                    else:
                        from core.validators import validate_birth_date
                        try:
                            validate_birth_date(birth_date_obj)
                        except ValidationError as e:
                            row_errors.append({"field": col_birth, "message": format_validation_error(e)})

            if col_join:
                join_str = self._clean(row_dict.get(col_join, ''))
                if join_str:
                    join_date_obj = self._parse_date(join_str)
                    if not join_date_obj:
                        row_errors.append({"field": col_join, "message": "تاريخ الالتحاق غير صالح"})
                    else:
                        from core.validators import validate_join_date
                        try:
                            validate_join_date(join_date_obj, birth_date_obj)
                        except ValidationError as e:
                            row_errors.append({"field": col_join, "message": format_validation_error(e)})

            # 4. الرتبة الصارمة
            raw_rank = self._clean(row_dict.get(col_rank, ''))
            rank_obj = None
            if raw_rank:
                mapped_rank = RANK_MAP.get(raw_rank, raw_rank)
                if mapped_rank not in self.ranks:
                    row_errors.append({"field": col_rank, "message": f"الرتبة غير معروفة في النظام: ({raw_rank})"})
                else:
                    rank_obj = self.ranks[mapped_rank]

            # 5. الحالة ونوع الحالة والترابط بينهما (بتسمية العميل: الحالة = التصنيف 4 خيارات، نوع الحالة = الحالة التفصيلية)
            raw_status = self._clean(row_dict.get(col_status, ''))
            raw_status_type = self._clean(row_dict.get(col_status_type, ''))
            
            valid_status_type = True
            valid_status = True
            
            # فحص الحالة (التصنيف العام - 4 خيارات)
            if raw_status:
                allowed_classifications = ['قوة عاملة فعلية', 'قوة عاملة غير فعلية', 'قوة غير عاملة مؤقتاً', 'قوة غير عاملة نهائياً']
                if raw_status not in allowed_classifications:
                    row_errors.append({"field": col_status, "message": f"الحالة (التصنيف العام) غير معروفة: ({raw_status})"})
                    valid_status = False

            # فحص نوع الحالة (التفصيلية)
            status_obj = None
            if raw_status_type:
                mapped_status = STATUS_MAP.get(raw_status_type, raw_status_type)
                if mapped_status not in self.statuses:
                    row_errors.append({"field": col_status_type, "message": f"نوع الحالة (التفصيلية) غير معروف في النظام: ({raw_status_type})"})
                    valid_status_type = False
                else:
                    status_obj = self.statuses[mapped_status]
            
            # فحص الترابط
            if valid_status_type and valid_status and raw_status and status_obj:
                classification_display = str(status_obj.get_classification_display())
                if classification_display != raw_status:
                    row_errors.append({
                        "field": col_status_type, 
                        "message": f"نوع الحالة ({raw_status_type}) لا ينتمي للحالة ({raw_status}). التصنيف الصحيح له هو ({classification_display})."
                    })

            # 6. تصنيف القوة والفئة والمنصب
            raw_force = self._clean(row_dict.get('تصنيف القوة', ''))
            force_obj = None
            if raw_force:
                mapped_force = FORCE_TYPE_MAP.get(raw_force, raw_force)
                if mapped_force not in self.force_types:
                    row_errors.append({"field": "تصنيف القوة", "message": f"تصنيف القوة غير معروف: ({raw_force})"})
                else:
                    force_obj = self.force_types[mapped_force]
                    
            # فحص الرتبة مقابل تصنيف القوة
            if rank_obj and force_obj:
                if force_obj.rank_type == 'officer' and not rank_obj.is_officer:
                    row_errors.append({
                        "field": col_rank,
                        "message": f"تصنيف القوة ({raw_force}) مخصص للضباط فقط، بينما الرتبة ({raw_rank}) للأفراد/الصف."
                    })
                elif force_obj.rank_type == 'personnel' and rank_obj.is_officer:
                    row_errors.append({
                        "field": col_rank,
                        "message": f"تصنيف القوة ({raw_force}) مخصص للأفراد والصف فقط، بينما الرتبة ({raw_rank}) لضابط."
                    })
            
            qual = self._clean(row_dict.get('المؤهل', ''))
            if qual and qual not in self.qualifications:
                row_errors.append({"field": "المؤهل", "message": f"المؤهل غير معروف في النظام: ({qual})"})
                
            cat = self._clean(row_dict.get('الفئة', ''))
            cat_obj = None
            if cat:
                if cat not in self.job_categories:
                    row_errors.append({"field": "الفئة", "message": f"الفئة غير معروفة في النظام: ({cat})"})
                else:
                    cat_obj = self.job_categories[cat]
                
            job = self._clean(row_dict.get('نوع العمل', ''))
            job_obj = None
            if job:
                if job not in self.job_titles:
                    row_errors.append({"field": "نوع العمل", "message": f"نوع العمل غير معروف: ({job})"})
                else:
                    job_obj = self.job_titles[job]
                
            pos = self._clean(row_dict.get('المنصب', ''))
            pos_obj = None
            if pos:
                if pos not in self.positions:
                    row_errors.append({"field": "المنصب", "message": f"المنصب غير معروف: ({pos})"})
                else:
                    pos_obj = self.positions[pos]

            # الترابط الذكي بين الفئة ونوع العمل والمنصب
            if job_obj and cat:
                if job_obj.category.name != cat:
                    row_errors.append({
                        "field": "نوع العمل", 
                        "message": f"نوع العمل ({job}) لا يتبع للفئة ({cat}). الفئة الصحيحة له هي ({job_obj.category.name})."
                    })
            
            if pos_obj:
                if cat:
                    allowed = pos_obj.allowed_categories
                    if allowed and isinstance(allowed, list) and cat not in allowed:
                        allowed_str = "، ".join(allowed)
                        row_errors.append({
                            "field": "المنصب", 
                            "message": f"المنصب ({pos}) غير مسموح للفئة ({cat}). الفئات المسموحة هي: ({allowed_str})."
                        })
                
                # فحص الرتبة المطلوبة للمنصب
                if rank_obj and pos_obj.requires_rank_id:
                    required_rank_obj = self.ranks.get(pos_obj.requires_rank.name)
                    if required_rank_obj and rank_obj.order > required_rank_obj.order:
                        row_errors.append({
                            "field": col_rank,
                            "message": f"المنصب ({pos}) يتطلب رتبة ({required_rank_obj.name}) كحد أدنى، ولكن الموظف برتبة ({rank_obj.name})."
                        })

            # 7. الهيكل التنظيمي الصارم
            unit_str = self._clean(row_dict.get(col_unit, ''))
            
            # محاولة تصحيح اسم الوحدة (مثل 'المحافظة الرئيسية' -> 'شرطة محافظة المحافظة الرئيسية')
            if unit_str in UNIT_MAP:
                unit_str = self._clean(UNIT_MAP[unit_str])
            elif row_dict.get(col_unit, '').strip() in UNIT_MAP:
                unit_str = self._clean(UNIT_MAP[row_dict.get(col_unit, '').strip()])
                
            unit_obj = None
            if unit_str:
                if unit_str not in self.security_admins:
                    row_errors.append({"field": col_unit, "message": f"الوحدة (الإدارة الأمنية) غير معروفة: ({row_dict.get(col_unit, '')})"})
                else:
                    unit_obj = self.security_admins[unit_str]

            dept_str = self._clean(row_dict.get(col_dept, ''))
            dept_obj = None
            dept_type = None
            
            if dept_str:
                found = False
                
                if dept_str in self.central_depts:
                    depts = self.central_depts[dept_str]
                    match = next((d for d in depts if unit_obj and d.security_admin_id == unit_obj.id), depts[0])
                    dept_obj = match
                    dept_type = 'إدارة مركزية'
                    found = True
                elif dept_str in self.branches:
                    depts = self.branches[dept_str]
                    match = next((d for d in depts if unit_obj and d.security_admin_id == unit_obj.id), depts[0])
                    dept_obj = match
                    dept_type = 'فرع'
                    found = True
                elif dept_str in self.district_police:
                    depts = self.district_police[dept_str]
                    match = next((d for d in depts if unit_obj and d.security_admin_id == unit_obj.id), depts[0])
                    dept_obj = match
                    dept_type = 'أمن مديرية'
                    found = True
                
                if not found:
                    row_errors.append({"field": col_dept, "message": f"جهة العمل (إدارة/فرع/مديرية) غير معروفة: ({dept_str})"})
            
            # فحص ارتباط جهة العمل بالوحدة (محافظة)
            if unit_obj and dept_obj:
                if dept_obj.security_admin_id != unit_obj.id:
                    row_errors.append({
                        "field": col_dept, 
                        "message": f"جهة العمل ({dept_str}) لا تتبع للوحدة ({unit_str})."
                    })

            div_str = self._clean(row_dict.get(col_div, ''))
            div_obj = None
            if div_str:
                if div_str not in self.divisions:
                    row_errors.append({"field": col_div, "message": f"القسم/فرع السرية غير معروف نهائياً في النظام: ({div_str})"})
                else:
                    divs = self.divisions[div_str]
                    # Find the specific division that belongs to the selected dept_obj
                    if dept_obj:
                        div_obj = next((d for d in divs if (d.central_department_id == dept_obj.id or d.branch_id == dept_obj.id or d.district_police_id == dept_obj.id)), divs[0])
                    else:
                        div_obj = divs[0]
            
            # فحص ارتباط القسم بجهة العمل
            if dept_obj and div_obj:
                parent_id = div_obj.central_department_id or div_obj.branch_id or div_obj.district_police_id
                if parent_id != dept_obj.id:
                    row_errors.append({
                        "field": col_div, 
                        "message": f"القسم ({div_str}) لا يتبع لجهة العمل ({dept_str})."
                    })

            # فحص حالة النفقات
            expense_status = self._clean(row_dict.get('حالة النفقات', ''))
            if expense_status and expense_status not in ['لديه نفقات', 'بدون نفقات']:
                row_errors.append({
                    "field": "حالة النفقات", 
                    "message": f"حالة النفقات غير معروفة: ({expense_status}). المسموح: 'لديه نفقات' أو 'بدون نفقات'."
                })

            if row_errors:
                errors.append({
                    "row": idx,
                    "military_number": mil_num,
                    "name": full_name,
                    "errors": row_errors,
                    "rowData": {k: self._clean(v) for k, v in row_dict.items()}
                })
            else:
                valid_rows += 1

        # فحص أرقام عسكرية مكررة في قاعدة البيانات إذا لم تكن هناك أخطاء أولية
        if not errors:
            existing_mil_nums = set(PersonnelMaster.objects.filter(
                military_number__in=file_military_numbers
            ).values_list('military_number', flat=True))
            
            if existing_mil_nums:
                for idx, row_dict in enumerate(rows_data, start=2):
                    mil_num = self._clean(row_dict.get(col_mil, ''))
                    if mil_num in existing_mil_nums:
                        errors.append({
                            "row": idx,
                            "military_number": mil_num,
                            "name": self._clean(row_dict.get(col_name, '')),
                            "errors": [{"field": col_mil, "message": f"الرقم العسكري ({mil_num}) مسجل مسبقاً في النظام"}],
                            "rowData": {k: self._clean(v) for k, v in row_dict.items()}
                        })
                        valid_rows -= 1

        # تحويل جميع الصفوف إلى strings للواجهة الأمامية
        all_rows_clean = [{k: self._clean(v) for k, v in row.items()} for row in rows_data]

        return {
            "total_rows": len(rows_data),
            "valid_rows": valid_rows,
            "error_count": len(errors),
            "is_valid": len(errors) == 0,
            "errors": errors,
            "all_rows": all_rows_clean
        }

    @transaction.atomic
    def commit_file(self, file_content: bytes, batch_id: str = None, column_mapping: Dict[str, str] = None) -> Dict[str, Any]:
        """
        الاعتماد النهائي من الإكسل مباشرة (يقرأ الملف ويحيله إلى commit_data).
        """
        rows_data = self.parse_file_to_dicts(file_content, column_mapping)
        return self.commit_data(rows_data, batch_id)

    @transaction.atomic
    def commit_data(self, rows_data: List[Dict[str, Any]], batch_id: str = None) -> Dict[str, Any]:
        """
        الاعتماد النهائي وحفظ البيانات من قائمة JSON (يجب أن يتم استدعاؤه فقط إذا تم تجاوز Validation بـ 0 أخطاء).
        """
        if not batch_id:
            batch_id = str(uuid.uuid4())
            
        if not rows_data:
            raise StrictImportError("لا توجد بيانات للاعتماد")
            
        headers = list(rows_data[0].keys())
        
        stats = {'created': 0, 'name_corrections': 0, 'monthly_vars': 0}
        
        col_mil = 'الرقم العسكري' if 'الرقم العسكري' in headers else ('الرقم العسكري القديم' if 'الرقم العسكري القديم' in headers else '')
        col_name = 'الاسم' if 'الاسم' in headers else ('الاسم الكامل' if 'الاسم الكامل' in headers else 'الأسم')
        col_name_correction = 'تصحيح الاسم من واقع البطاقة'
        
        monthly_cols = [h for h in headers if h.startswith('متغير')]

        for row_idx, row_dict in enumerate(rows_data, start=2):
            # حفظ النسخة الخام للأرشيف
            raw_data = {k: str(v) if v is not None else '' for k, v in row_dict.items()}
            RawDataImport.objects.create(
                row_index=row_idx,
                raw_data=raw_data,
                import_batch_id=batch_id,
                status='processed'
            )

            # البيانات الأساسية
            mil_correct = self._clean(row_dict.get('الرقم العسكري الصحيح', ''))
            mil_normal = self._clean(row_dict.get('الرقم العسكري', ''))
            mil_old = self._clean(row_dict.get('الرقم العسكري القديم', ''))
            mil_num = mil_normal or mil_correct or mil_old
            
            full_name = self._clean(row_dict.get(col_name, ''))
            nat_id = self._clean(row_dict.get('الرقم الوطني', '')) or None
            
            raw_rank = self._clean(row_dict.get('الرتبة', ''))
            rank = self.ranks.get(RANK_MAP.get(raw_rank, raw_rank))
            
            raw_status = self._clean(row_dict.get('الحالة', '') or row_dict.get('نوع الحالة', ''))
            status_obj = self.statuses.get(STATUS_MAP.get(raw_status, raw_status))
            
            raw_force = self._clean(row_dict.get('تصنيف القوة', ''))
            force = self.force_types.get(FORCE_TYPE_MAP.get(raw_force, raw_force))
            if not force and mil_num:
                # استنتاج تلقائي كما كان في السكربت القديم إذا كانت فارغة
                if mil_num[:2] == '50':
                    force = self.force_types.get('لجان أمنية - أفراد')
                elif mil_num[:2] == '60':
                    force = self.force_types.get('القوة الأساسي - ضباط')
                elif mil_num[:1] == '7':
                    force = self.force_types.get('القوة الأساسي - أفراد')
            
            qual = self.qualifications.get(self._clean(row_dict.get('المؤهل', '')))
            cat = self.job_categories.get(self._clean(row_dict.get('الفئة', '')))
            job = self.job_titles.get(self._clean(row_dict.get('نوع العمل', '')))
            pos = self.positions.get(self._clean(row_dict.get('المنصب', '')))
            
            # تطبيق UNIT_MAP قبل استخراج الإدارة الأمنية
            raw_unit = self._clean(row_dict.get('الوحدة', ''))
            if raw_unit in UNIT_MAP:
                raw_unit = self._clean(UNIT_MAP[raw_unit])
            elif row_dict.get('الوحدة', '').strip() in UNIT_MAP:
                raw_unit = self._clean(UNIT_MAP[row_dict.get('الوحدة', '').strip()])
            
            sec_admin = self.security_admins.get(raw_unit)
            
            c_dept = self.central_depts.get(self._clean(row_dict.get('الإدارة_السرية', '')))
            
            div_str = self._clean(row_dict.get('القسم_فرع السرية', ''))
            branch = self.branches.get(div_str)
            dist_p = None if branch else self.district_police.get(div_str)
            div = None if (branch or dist_p) else self.divisions.get(div_str)
            
            birth_date = self._parse_date(row_dict.get('تاريخ الميلاد'))
            join_date = self._parse_date(row_dict.get('تاريخ الألتحاق'))
            dec_date = self._parse_date(row_dict.get('تاريخ صدور القرار'))
            trans_date = self._parse_date(row_dict.get('تاريخ التصدور الينا'))

            raw_expense = self._clean(row_dict.get('حالة النفقات', ''))
            expense_val = None
            for k, v in EXPENSE_MAP.items():
                if k in raw_expense:
                    expense_val = v
                    break

            notes = self._clean(row_dict.get('الملاحظات', ''))
            appointment_info = self._clean(row_dict.get('التعيينات', '')) or None

            # إنشاء الفرد
            person = PersonnelMaster.objects.create(
                military_number=mil_num,
                full_name=full_name,
                national_id=nat_id,
                old_military_number=mil_old or None,
                current_rank=rank,
                current_status=status_obj,
                force_classification=force,
                qualification=qual,
                category=cat,
                job_title=job,
                position=pos,
                security_admin=sec_admin,
                central_department=c_dept,
                branch=branch,
                district_police=dist_p,
                division=div,
                birth_date=birth_date,
                join_date=join_date,
                decision_date=dec_date,
                transfer_date=trans_date,
                phone_number=self._clean(row_dict.get('رقم التليفون', '')) or None,
                expense_status=expense_val,
                appointment_info=appointment_info,
                notes=notes,
                is_data_clean=True, # تم التحقق بصرامة
                data_quality_score=100
            )
            stats['created'] += 1

            # مقترح تصحيح الاسم
            if col_name_correction in headers:
                corr_name = self._clean(row_dict.get(col_name_correction, ''))
                if corr_name and corr_name != full_name:
                    SuggestedCorrection.objects.create(
                        personnel=person,
                        security_admin=person.security_admin,
                        correction_type='name_correction',
                        field_name='full_name',
                        old_value=full_name,
                        new_value=corr_name,
                        status='pending',
                        requested_at=timezone.now(),
                    )
                    stats['name_corrections'] += 1

            # المتغيرات الشهرية
            for m_col in monthly_cols:
                val = self._clean(row_dict.get(m_col, ''))
                if val:
                    HistoricalMonthlyVariables.objects.create(
                        personnel=person,
                        month=m_col,
                        variable_value=val,
                        source_column=m_col
                    )
                    stats['monthly_vars'] += 1
                    
        return stats
