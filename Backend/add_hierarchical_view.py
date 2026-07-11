import os

file_path = 'systems/personnel/api/views/reports_views.py'

new_view = """
from openpyxl import Workbook
from django.http import HttpResponse

class HierarchicalWorkforceView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        export = request.query_params.get('export', 'false') == 'true'
        
        from infra.authorization.services.permission_service import PermissionService
        qs = PermissionService.get_scoped_queryset(
            request.user, PersonnelMaster.objects.all(), 'personnel.view.*'
        )
        
        # 1. Aggregate down to the Unit level
        aggs = qs.values(
            'security_admin__name',
            'central_department__name',
            'branch__name',
            'district_police__name',
            'division__name',
            'unit__name',
            'current_rank__name'
        ).annotate(count=Count('military_number'))
        
        officer_ranks = [
            "فريق أول", "فريق", "لواء", "عميد", "عقيد", "مقدم", "رائد", "نقيب", "ملازم أول", "ملازم ثاني"
        ]
        
        # 2. Build the Tree
        tree = {}
        
        for row in aggs:
            sa = row['security_admin__name'] or 'إدارة غير محددة'
            
            # Determine Level 1 (Dept/Branch/District)
            l1 = row['central_department__name'] or row['branch__name'] or row['district_police__name'] or 'جهة غير محددة'
            
            # Determine Level 2 & 3
            div = row['division__name'] or 'أقسام عامة'
            unit = row['unit__name'] or 'وحدات عامة'
            
            rank = row['current_rank__name'] or 'غير محدد'
            count = row['count']
            
            is_officer = rank in officer_ranks
            
            if sa not in tree:
                tree[sa] = {'name': sa, 'type': 'sa', 'officers': 0, 'ncos': 0, 'total': 0, 'children': {}}
            
            if l1 not in tree[sa]['children']:
                tree[sa]['children'][l1] = {'name': l1, 'type': 'l1', 'officers': 0, 'ncos': 0, 'total': 0, 'children': {}}
                
            if div not in tree[sa]['children'][l1]['children']:
                tree[sa]['children'][l1]['children'][div] = {'name': div, 'type': 'div', 'officers': 0, 'ncos': 0, 'total': 0, 'children': {}}
                
            if unit not in tree[sa]['children'][l1]['children'][div]['children']:
                tree[sa]['children'][l1]['children'][div]['children'][unit] = {'name': unit, 'type': 'unit', 'officers': 0, 'ncos': 0, 'total': 0, 'ranks': {}}
                
            # Add counts
            node_unit = tree[sa]['children'][l1]['children'][div]['children'][unit]
            node_unit['total'] += count
            if is_officer: node_unit['officers'] += count
            else: node_unit['ncos'] += count
            node_unit['ranks'][rank] = node_unit['ranks'].get(rank, 0) + count
            
            # Propagate up
            for node in [tree[sa], tree[sa]['children'][l1], tree[sa]['children'][l1]['children'][div]]:
                node['total'] += count
                if is_officer: node['officers'] += count
                else: node['ncos'] += count
        
        # 3. Format as a list of nested dicts instead of keyed dicts for easier frontend rendering
        def dict_to_list(d):
            lst = []
            for k, v in d.items():
                if 'children' in v:
                    v['children'] = dict_to_list(v['children'])
                lst.append(v)
            # Sort by total descending
            return sorted(lst, key=lambda x: x['total'], reverse=True)
            
        structured_tree = dict_to_list(tree)
        
        if export:
            return self._export_excel(structured_tree)
            
        return Response({"data": structured_tree})
        
    def _export_excel(self, tree_data):
        wb = Workbook()
        ws = wb.active
        ws.title = "الهيكل التنظيمي للقوة"
        ws.rightToLeft = True
        
        # Headers
        headers = ['المستوى', 'الجهة / القسم', 'الضباط', 'الأفراد', 'الإجمالي']
        ws.append(headers)
        
        # Styling
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            
        # Recursive function to write rows
        def write_node(node, indent=0):
            indent_str = "    " * indent
            prefix = "▶ " if indent == 0 else ("  ▼ " if indent == 1 else ("    - " if indent == 2 else "      • "))
            row = [
                node.get('type', ''),
                f"{prefix}{node['name']}",
                node['officers'],
                node['ncos'],
                node['total']
            ]
            ws.append(row)
            
            # Style based on indent
            current_row = ws.max_row
            if indent == 0:
                for cell in ws[current_row]: cell.font = Font(bold=True, color='B91C1C')
            elif indent == 1:
                for cell in ws[current_row]: cell.font = Font(bold=True, color='0F766E')
                
            for child in node.get('children', []):
                write_node(child, indent + 1)
                
        for root_node in tree_data:
            write_node(root_node)
            
        ws.column_dimensions['B'].width = 50
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="workforce_tree.xlsx"'
        wb.save(response)
        return response
"""

with open(file_path, 'a', encoding='utf-8') as f:
    f.write(new_view)

print("View added successfully!")
