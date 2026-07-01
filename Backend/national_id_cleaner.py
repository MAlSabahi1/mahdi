import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==========================================
# 📍 إعدادات مسار الملفات (قم بتعديل المسار هنا)
# ==========================================
# ضع مسار ملف الإكسل الخام الخاص بك هنا
INPUT_FILE_PATH = '/home/mahdi/Desktop/خدمات أمن المحافظة الأولى يناير 2026 -م.xlsx'

# مسار الملف الناتج بعد التنظيف
OUTPUT_FILE_PATH = "/home/mahdi/Desktop/POL/cleaned_national_ids.xlsx"

# رقم الورقة (Sheet) - الورقة الثانية تعني الفهرس 1 (لأن العد يبدأ من 0)
SHEET_INDEX = 1 
# ==========================================

# تعريف الألوان
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # أصفر
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # أحمر

def clean_nid(val):
    """دالة مساعدة لتنظيف القيم من المسافات والفواصل"""
    if pd.isna(val) or str(val).strip().lower() in ['nan', 'none', '']:
        return ""
    val = str(val).strip().replace(".0", "").replace(" ", "").replace("-", "")
    return val

def process_data():
    print(f"جاري قراءة البيانات من الورقة الثانية في الملف: {INPUT_FILE_PATH} ...")
    try:
        # قراءة الورقة الثانية (sheet_name=1)
        df = pd.read_excel(INPUT_FILE_PATH, sheet_name=SHEET_INDEX, dtype=str)
    except Exception as e:
        print(f"حدث خطأ أثناء قراءة الملف. تأكد من المسار واسم الورقة: {e}")
        return

    # إنشاء عمود جديد للنتيجة النهائية
    df['الرقم الوطني المعتمد'] = ""
    
    # قاموس لتتبع أرقام الصفوف التي تحتاج تلوين
    color_actions = {} 

    print("جاري المطابقة وتحليل الرقم الوطني...")
    
    for i, row in df.iterrows():
        nid = clean_nid(row.get('الرقم الوطني'))
        cnid = clean_nid(row.get('تصحيح الرقم الوطني'))
        
        # 1. إذا كان العمودين فارغين
        if not nid and not cnid:
            df.at[i, 'الرقم الوطني المعتمد'] = ""
            
        # 2. يوجد رقم وطني ولكن التصحيح فارغ
        elif nid and not cnid:
            # نعتمد الرقم الوطني الأساسي
            # نضيف له صفر في البداية إذا كان 10 أرقام ليصبح 11
            df.at[i, 'الرقم الوطني المعتمد'] = nid.zfill(11)
            
        # 3. الرقم الوطني فارغ والتصحيح موجود
        elif not nid and cnid:
            df.at[i, 'الرقم الوطني المعتمد'] = cnid.zfill(11)
            color_actions[i] = "YELLOW" # تحديد للتلوين بالأصفر
            
        # 4. الرقمان موجودان (يجب المطابقة)
        elif nid and cnid:
            # نقوم بإضافة الأصفار الناقصة (zfill) لكلا الرقمين لمطابقتهما 
            # لمعالجة مشكلة "ينقصه الصفر في البداية"
            if nid.zfill(11) == cnid.zfill(11):
                # متطابقان تماماً أو متطابقان بعد إضافة الصفر المفقود
                df.at[i, 'الرقم الوطني المعتمد'] = cnid.zfill(11)
            else:
                # مختلفان تماماً
                df.at[i, 'الرقم الوطني المعتمد'] = cnid # نضع رقم التصحيح
                color_actions[i] = "RED" # تحديد للتلوين بالأحمر

    print("جاري حفظ الملف الأولي...")
    # حفظ الملف بصيغة إكسل أولاً
    df.to_excel(OUTPUT_FILE_PATH, index=False)
    
    # استخدام openpyxl لتطبيق الألوان على الخلايا
    print("جاري تلوين الخلايا (الأصفر والأحمر)...")
    wb = load_workbook(OUTPUT_FILE_PATH)
    ws = wb.active
    
    # البحث عن رقم عمود "الرقم الوطني المعتمد" و "تصحيح الرقم الوطني"
    approved_col_idx = df.columns.get_loc('الرقم الوطني المعتمد') + 1
    cnid_col_idx = df.columns.get_loc('تصحيح الرقم الوطني') + 1 if 'تصحيح الرقم الوطني' in df.columns else None
    
    for row_idx, color in color_actions.items():
        # رقم الصف في إكسل (row_idx + 2) 
        # لأن pandas index يبدأ من 0 والصف الأول في إكسل للعناوين
        excel_row = row_idx + 2
        
        # تلوين خلية الرقم المعتمد
        cell_approved = ws.cell(row=excel_row, column=approved_col_idx)
        
        # تلوين خلية التصحيح أيضاً لكي تكون واضحة جداً كما طلبت
        if cnid_col_idx:
            cell_cnid = ws.cell(row=excel_row, column=cnid_col_idx)

        if color == "YELLOW":
            cell_approved.fill = YELLOW_FILL
            if cnid_col_idx: cell_cnid.fill = YELLOW_FILL
        elif color == "RED":
            cell_approved.fill = RED_FILL
            if cnid_col_idx: cell_cnid.fill = RED_FILL

    wb.save(OUTPUT_FILE_PATH)
    print(f"تم الانتهاء بنجاح! تم حفظ الملف النهائي الملون في: {OUTPUT_FILE_PATH}")

if __name__ == "__main__":
    process_data()
