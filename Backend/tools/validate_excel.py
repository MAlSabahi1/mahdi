#!/usr/bin/env python3
"""
أداة فحص وتنظيف بيانات الإكسل — الأعمدة العسكرية
═══════════════════════════════════════════════════
يفحص: الرقم العسكري، الرقم القديم، رقم الضابط، الرتبة، الرتبة الجديدة، تصنيف القوة
المرجع: الدليل الإرشادي + قواعد النظام (personnel/models.py)

الاستخدام:
    python validate_excel.py /path/to/file.xlsx
    python validate_excel.py /path/to/file.xlsx --sheet "اسم الورقة"
    python validate_excel.py /path/to/file.xlsx --fix --output cleaned.xlsx
"""
import sys
import re
import argparse
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("❌ مطلوب: pip install openpyxl")
    sys.exit(1)

# ═══════════════════════════════════════════════════
# القواعد — مأخوذة من models.py والدليل الإرشادي
# ═══════════════════════════════════════════════════

OFFICER_RANKS = {'لواء', 'عميد', 'عقيد', 'مقدم', 'رائد', 'نقيب', 'ملازم أول', 'ملازم ثاني'}
PERSONNEL_RANKS = {
    'مساعد 1', 'مساعد 2', 'مساعد', 'رقيب 1', 'رقيب 2',
    'عريف', 'جندي', 'مجند', 'حارس', 'مدني', 'متعاقد',
    # أشكال كتابة بديلة
    'مساعد1', 'مساعد2', 'رقيب1', 'رقيب2',
}
ALL_RANKS = OFFICER_RANKS | PERSONNEL_RANKS

# ترتيب الرتب (order أقل = رتبة أعلى)
RANK_ORDER = {
    'لواء': 1, 'عميد': 2, 'عقيد': 3, 'مقدم': 4, 'رائد': 5,
    'نقيب': 6, 'ملازم أول': 7, 'ملازم ثاني': 8,
    'مساعد 1': 9, 'مساعد1': 9, 'مساعد 2': 10, 'مساعد2': 10,
    'مساعد': 11, 'رقيب 1': 13, 'رقيب1': 13, 'رقيب 2': 14, 'رقيب2': 14,
    'عريف': 15, 'جندي': 16, 'مجند': 17, 'حارس': 18, 'مدني': 19, 'متعاقد': 20,
}

FORCE_OFFICER = {'القوة الأساسي - ضباط', 'القوة الأساسي- ضباط', 'لجان أمنية - ضباط', 'لجان أمنية- ضباط'}
FORCE_PERSONNEL = {
    'القوة الأساسي - أفراد', 'القوة الأساسي- أفراد',
    'لجان أمنية - أفراد', 'لجان أمنية- أفراد',
    'قوة - مستجدين جديد', 'قوة- مستجدين جديد',
    'قوة إداريا فقط', 'إداريا فقط',
    'قوة منزلة - خارج الصرف', 'قوة منزلة- خارج الصرف',
}

# أسماء الأعمدة المتوقعة
COL_MIL = 'الرقم العسكري'
COL_MIL_CORRECT = 'الرقم العسكري الصحيح'
COL_MIL_OLD = 'الرقم العسكري القديم'
COL_OFFICER = 'رقم الضابط'
COL_RANK = 'الرتبة'
COL_RANK_NEW = 'الرتبة الجديدة'
COL_FORCE = 'تصنيف القوة'
COL_NAME = 'الأسم'
COL_NATIONAL = 'الرقم الوطني'

COLORS = {
    'RED': '\033[91m', 'GREEN': '\033[92m', 'YELLOW': '\033[93m',
    'BLUE': '\033[94m', 'CYAN': '\033[96m', 'BOLD': '\033[1m',
    'END': '\033[0m',
}

def c(text, color):
    return f"{COLORS.get(color, '')}{text}{COLORS['END']}"


def clean(val):
    if val is None:
        return ''
    return str(val).strip()


def is_officer_prefix(mil_num):
    return mil_num[:2] == '60' if len(mil_num) >= 2 else False

def is_basic_prefix(mil_num):
    return mil_num[0] == '7' if len(mil_num) >= 1 else False

def is_committee_prefix(mil_num):
    return mil_num[0] == '5' if len(mil_num) >= 1 else False

def get_prefix_type(mil_num):
    if is_officer_prefix(mil_num): return 'ضابط'
    if is_basic_prefix(mil_num): return 'أساسي أفراد'
    if is_committee_prefix(mil_num): return 'لجان/مستجدين'
    return 'غير معروف'

def is_officer_rank(rank):
    return rank in OFFICER_RANKS


# ═══════════════════════════════════════════════════
# محرك الفحص
# ═══════════════════════════════════════════════════

class ExcelValidator:
    def __init__(self, file_path, sheet_name=None):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active
        rows = list(self.ws.iter_rows(values_only=True))
        self.headers = [clean(h) for h in rows[0]] if rows else []
        self.data = rows[1:] if len(rows) > 1 else []
        self.issues = []  # (row_idx, severity, column, message, suggestion)
        self.stats = Counter()

    def _get(self, row_dict, col):
        return clean(row_dict.get(col, ''))

    def _add_issue(self, row, sev, col, msg, suggestion=''):
        self.issues.append((row, sev, col, msg, suggestion))
        self.stats[sev] += 1

    def validate(self):
        print(c(f"\n{'═'*60}", 'CYAN'))
        print(c("  أداة فحص بيانات الإكسل — الأعمدة العسكرية", 'BOLD'))
        print(c(f"{'═'*60}\n", 'CYAN'))
        print(f"📂 الملف: {self.file_path}")
        print(f"📊 الأعمدة: {len(self.headers)} | الصفوف: {len(self.data)}")

        # تحقق من وجود الأعمدة المطلوبة
        missing = []
        for col in [COL_MIL, COL_RANK]:
            if col not in self.headers:
                missing.append(col)
        if missing:
            print(c(f"\n❌ أعمدة مفقودة: {', '.join(missing)}", 'RED'))
            return

        found_cols = [h for h in self.headers if h in
                      [COL_MIL, COL_MIL_CORRECT, COL_MIL_OLD, COL_OFFICER,
                       COL_RANK, COL_RANK_NEW, COL_FORCE, COL_NAME, COL_NATIONAL]]
        print(f"🔍 الأعمدة المفحوصة: {', '.join(found_cols)}\n")

        mil_seen = defaultdict(list)  # لكشف التكرار

        for i, row in enumerate(self.data):
            row_dict = dict(zip(self.headers, row))
            row_num = i + 2  # +2 لأن الصف 1 = عناوين

            mil = self._get(row_dict, COL_MIL)
            mil_correct = self._get(row_dict, COL_MIL_CORRECT)
            mil_old = self._get(row_dict, COL_MIL_OLD)
            officer_num = self._get(row_dict, COL_OFFICER)
            rank = self._get(row_dict, COL_RANK)
            rank_new = self._get(row_dict, COL_RANK_NEW)
            force = self._get(row_dict, COL_FORCE)
            name = self._get(row_dict, COL_NAME)

            # ─── 1. الرقم العسكري ───
            effective_mil = mil_correct if mil_correct else mil
            if not effective_mil:
                self._add_issue(row_num, '🔴', COL_MIL, 'الرقم العسكري فارغ!')
                continue

            if not re.match(r'^\d{7}$', effective_mil):
                self._add_issue(row_num, '🔴', COL_MIL,
                    f'الرقم العسكري "{effective_mil}" ليس 7 أرقام',
                    f'تأكد أنه رقم من 7 خانات بالضبط')

            mil_seen[effective_mil].append(row_num)

            # ─── 2. الرقم العسكري القديم ───
            if mil_old:
                if not re.match(r'^\d{7}$', mil_old):
                    self._add_issue(row_num, '🔴', COL_MIL_OLD,
                        f'الرقم القديم "{mil_old}" ليس 7 أرقام')
                if mil_old == effective_mil:
                    self._add_issue(row_num, '🔴', COL_MIL_OLD,
                        'الرقم القديم = الرقم الحالي! يجب أن يكونا مختلفين',
                        'احذف الرقم القديم أو صحح أحدهما')

            # ─── 3. رقم الضابط (يجب حذفه أو معالجته) ───
            if officer_num:
                if officer_num == effective_mil:
                    self._add_issue(row_num, '⚠️', COL_OFFICER,
                        f'رقم الضابط = الرقم العسكري → يمكن حذفه (مُكرر)',
                        'سيتم حذف عمود رقم الضابط من النظام')
                elif re.match(r'^60\d{5}$', officer_num):
                    if not is_officer_prefix(effective_mil):
                        self._add_issue(row_num, '🟡', COL_OFFICER,
                            f'رقم الضابط ({officer_num}) موجود لكن الرقم العسكري ({effective_mil}) ليس ضابطاً',
                            f'هذا فرد يحتاج تسوية: الرقم الجديد={officer_num}, القديم={effective_mil}')

            # ─── 4. الرتبة ───
            if rank and rank not in ALL_RANKS:
                self._add_issue(row_num, '🔴', COL_RANK,
                    f'رتبة غير معروفة: "{rank}"',
                    f'الرتب المسموحة: {", ".join(sorted(ALL_RANKS))}')

            # ─── 5. توافق بادئة الرقم مع الرتبة ───
            if rank and effective_mil and re.match(r'^\d{7}$', effective_mil):
                prefix_type = get_prefix_type(effective_mil)
                rank_is_officer = is_officer_rank(rank)

                if is_officer_prefix(effective_mil) and not rank_is_officer:
                    self._add_issue(row_num, '🔴', COL_RANK,
                        f'بادئة 60 (ضابط) لكن الرتبة "{rank}" هي رتبة أفراد!',
                        f'غيّر الرتبة لرتبة ضباط أو صحح الرقم العسكري')

                if (is_basic_prefix(effective_mil) or is_committee_prefix(effective_mil)) and rank_is_officer:
                    self._add_issue(row_num, '🔴', COL_RANK,
                        f'بادئة {effective_mil[:2]} (أفراد) لكن الرتبة "{rank}" هي رتبة ضباط!',
                        f'غيّر الرتبة لرتبة أفراد أو صحح الرقم العسكري')

            # ─── 6. الرتبة الجديدة ───
            if rank_new:
                if rank_new not in ALL_RANKS:
                    self._add_issue(row_num, '🔴', COL_RANK_NEW,
                        f'رتبة جديدة غير معروفة: "{rank_new}"')

                if rank and rank_new == rank:
                    self._add_issue(row_num, '⚠️', COL_RANK_NEW,
                        f'الرتبة الجديدة = الرتبة الحالية ({rank}) → لا حاجة للتسوية',
                        'احذف الرتبة الجديدة')

                if rank and rank_new and rank in RANK_ORDER and rank_new in RANK_ORDER:
                    if RANK_ORDER[rank_new] > RANK_ORDER[rank]:
                        self._add_issue(row_num, '🔴', COL_RANK_NEW,
                            f'الرتبة الجديدة "{rank_new}" أقل من الحالية "{rank}" — تخفيض!',
                            'المرجع: الدليل الإرشادي البند 17 — لا تخفيض')

                # تسوية فرد→ضابط: يجب وجود رقم ضابط يبدأ بـ 60
                if rank_new in OFFICER_RANKS and rank in PERSONNEL_RANKS:
                    if not officer_num or not officer_num.startswith('60'):
                        self._add_issue(row_num, '🟡', COL_RANK_NEW,
                            f'تسوية فرد→ضابط ({rank}→{rank_new}) لكن لا يوجد رقم ضابط يبدأ بـ 60',
                            'أدخل الرقم العسكري الجديد (يبدأ بـ 60) في عمود "الرقم العسكري الصحيح"')

            # ─── 7. تصنيف القوة ───
            if force and rank:
                rank_is_officer = is_officer_rank(rank)
                if force in FORCE_OFFICER and not rank_is_officer:
                    self._add_issue(row_num, '🔴', COL_FORCE,
                        f'تصنيف "{force}" مخصص للضباط لكن الرتبة "{rank}" أفراد!')
                if force in FORCE_PERSONNEL and rank_is_officer:
                    self._add_issue(row_num, '🔴', COL_FORCE,
                        f'تصنيف "{force}" مخصص للأفراد لكن الرتبة "{rank}" ضباط!')

                # توافق البادئة مع تصنيف القوة
                if effective_mil and re.match(r'^\d{7}$', effective_mil):
                    if is_basic_prefix(effective_mil) and 'لجان' in force:
                        self._add_issue(row_num, '⚠️', COL_FORCE,
                            f'بادئة 7 (أساسي) لكن التصنيف "{force}" لجان — تحقق!',
                            'بادئة 7 عادةً = "القوة الأساسي - أفراد"')

        # ─── كشف التكرارات ───
        for mil, rows in mil_seen.items():
            if len(rows) > 1:
                for r in rows:
                    self._add_issue(r, '🔴', COL_MIL,
                        f'الرقم العسكري {mil} مكرر في الصفوف: {rows}',
                        'احذف المكرر أو صحح الرقم')

        self._print_report()

    def _print_report(self):
        print(c(f"\n{'═'*60}", 'CYAN'))
        print(c("  📋 تقرير الفحص", 'BOLD'))
        print(c(f"{'═'*60}\n", 'CYAN'))

        total = len(self.issues)
        critical = self.stats.get('🔴', 0)
        warnings = self.stats.get('⚠️', 0)
        info = self.stats.get('🟡', 0)

        if total == 0:
            print(c("  ✅ لا توجد مشاكل! البيانات نظيفة وجاهزة للرفع.", 'GREEN'))
            return

        print(f"  إجمالي المشاكل: {c(total, 'BOLD')}")
        print(f"  🔴 حرج (يمنع الرفع): {c(critical, 'RED')}")
        print(f"  ⚠️ تحذير (يفضل إصلاحه): {c(warnings, 'YELLOW')}")
        print(f"  🟡 ملاحظة (للمراجعة): {c(info, 'BLUE')}")

        # تجميع حسب العمود
        by_col = defaultdict(list)
        for row, sev, col, msg, sug in self.issues:
            by_col[col].append((row, sev, msg, sug))

        for col, items in sorted(by_col.items()):
            print(c(f"\n── {col} ({len(items)} مشكلة) ──", 'BOLD'))
            for row, sev, msg, sug in items[:20]:  # أول 20
                print(f"  صف {row:>4} {sev} {msg}")
                if sug:
                    print(f"           💡 {c(sug, 'CYAN')}")
            if len(items) > 20:
                print(c(f"  ... و {len(items)-20} مشكلة أخرى", 'YELLOW'))

        # ملخص المكررات
        dups = [i for i in self.issues if 'مكرر' in i[3]]
        if dups:
            unique_mils = set()
            for _, _, _, msg, _ in dups:
                m = re.search(r'(\d{7})', msg)
                if m: unique_mils.add(m.group(1))
            print(c(f"\n⚠️ أرقام عسكرية مكررة: {len(unique_mils)} رقم", 'RED'))

    def fix_and_save(self, output_path):
        """إصلاح تلقائي للمشاكل القابلة للإصلاح وحفظ ملف جديد"""
        print(c(f"\n{'═'*60}", 'GREEN'))
        print(c("  🔧 الإصلاح التلقائي", 'BOLD'))
        print(c(f"{'═'*60}\n", 'GREEN'))

        fixes = 0
        for i, row in enumerate(self.data):
            row_dict = dict(zip(self.headers, row))
            row_num = i + 2
            row_list = list(row)

            mil = self._get(row_dict, COL_MIL)
            mil_correct = self._get(row_dict, COL_MIL_CORRECT)
            officer_num = self._get(row_dict, COL_OFFICER)
            rank = self._get(row_dict, COL_RANK)
            rank_new = self._get(row_dict, COL_RANK_NEW)

            # إصلاح 1: إذا الرقم الصحيح موجود → استبدل الرقم العسكري
            if mil_correct and mil_correct != mil and COL_MIL_CORRECT in self.headers:
                mil_idx = self.headers.index(COL_MIL)
                old_idx = self.headers.index(COL_MIL_OLD) if COL_MIL_OLD in self.headers else None
                correct_idx = self.headers.index(COL_MIL_CORRECT)

                # احفظ القديم
                if old_idx is not None and not self._get(row_dict, COL_MIL_OLD):
                    row_list[old_idx] = mil
                    print(f"  صف {row_num}: حفظ الرقم القديم {mil}")

                row_list[mil_idx] = mil_correct
                row_list[correct_idx] = ''  # فرّغ العمود بعد النقل
                print(f"  صف {row_num}: نقل الرقم الصحيح {mil_correct} → الرقم العسكري")
                fixes += 1

            # إصلاح 2: رقم الضابط = الرقم العسكري → احذفه
            effective = mil_correct if mil_correct else mil
            if officer_num and officer_num == effective and COL_OFFICER in self.headers:
                off_idx = self.headers.index(COL_OFFICER)
                row_list[off_idx] = ''
                print(f"  صف {row_num}: حذف رقم ضابط مكرر ({officer_num})")
                fixes += 1

            # إصلاح 3: رتبة جديدة = الرتبة الحالية → احذفها
            if rank_new and rank_new == rank and COL_RANK_NEW in self.headers:
                rn_idx = self.headers.index(COL_RANK_NEW)
                row_list[rn_idx] = ''
                print(f"  صف {row_num}: حذف رتبة جديدة مكررة ({rank_new})")
                fixes += 1

            # إصلاح 4: رقم ضابط يبدأ بـ 60 + فرد → نقل لعمود الرقم الصحيح
            if (officer_num and re.match(r'^60\d{5}$', officer_num)
                    and not is_officer_prefix(effective)
                    and rank_new and is_officer_rank(rank_new)):
                if COL_MIL_CORRECT in self.headers:
                    correct_idx = self.headers.index(COL_MIL_CORRECT)
                    if not row_list[correct_idx]:
                        row_list[correct_idx] = officer_num
                        print(f"  صف {row_num}: نقل رقم الضابط {officer_num} → الرقم الصحيح (تسوية)")
                        fixes += 1

            # تطبيق التعديلات
            self.data[i] = tuple(row_list)

        # حفظ
        wb_new = openpyxl.Workbook()
        ws_new = wb_new.active
        ws_new.append(self.headers)
        for row in self.data:
            ws_new.append(list(row))

        wb_new.save(output_path)
        print(c(f"\n✅ تم حفظ {output_path} — {fixes} إصلاح تلقائي", 'GREEN'))
        print(c("⚠️ أعد تشغيل الفحص على الملف الجديد للتأكد!", 'YELLOW'))


def main():
    parser = argparse.ArgumentParser(
        description='أداة فحص وتنظيف بيانات الإكسل — الأعمدة العسكرية',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
أمثلة:
  python validate_excel.py data.xlsx                    # فحص فقط
  python validate_excel.py data.xlsx --fix -o clean.xlsx  # فحص + إصلاح
  python validate_excel.py data.xlsx --sheet "المحافظة الأولى"     # ورقة محددة
        """
    )
    parser.add_argument('file', help='مسار ملف الإكسل')
    parser.add_argument('--sheet', '-s', help='اسم الورقة (افتراضي: الأولى)')
    parser.add_argument('--fix', action='store_true', help='تطبيق الإصلاحات التلقائية')
    parser.add_argument('--output', '-o', default='cleaned_output.xlsx', help='مسار الملف المُصلح')

    args = parser.parse_args()

    validator = ExcelValidator(args.file, args.sheet)
    validator.validate()

    if args.fix:
        validator.fix_and_save(args.output)


if __name__ == '__main__':
    main()
