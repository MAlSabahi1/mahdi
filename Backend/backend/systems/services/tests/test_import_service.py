"""
اختبارات خدمة الاستيراد - المهمة 2.2

تتحقق من:
- قراءة ملف Excel المعدل
- التحقق من UUIDs
- اكتشاف التغييرات
- تصنيف التغييرات (أخضر/أصفر/أحمر)
- إنشاء StagingRecord
- تقرير الأخطاء والتحذيرات
"""
import pytest
import uuid
from io import BytesIO
from django.contrib.auth import get_user_model
from django.utils import timezone
import openpyxl

from core.models import GeoGovernorate, SecurityAdministration, CentralDepartment, GeoDistrict, Rank, ServiceStatus, Qualification
from systems.personnel.models import PersonnelMaster
from systems.services.export_service import ExcelExportService
from systems.services.import_service import ExcelImportService, ImportValidationError, import_service_file
from systems.services.models import ExportLog, StagingRecord, DirectorateCompliance

User = get_user_model()


@pytest.fixture
def setup_test_data(db):
    """تجهيز بيانات الاختبار"""
    # إنشاء مستخدم
    user = User.objects.create_user(
        username='testuser',
        password='testpass123',
        is_staff=True
    )
    
    # إنشاء محافظة
    governorate = GeoGovernorate.objects.create(name_ar='البصرة')
    security_admin = SecurityAdministration.objects.create(name='أمن العاصمة', geo_governorate=governorate)

    # إنشاء إدارة
    department = CentralDepartment.objects.create(
        name='إدارة الاختبار',
        code='TEST001',
       
        security_admin=security_admin
    )
    
    # إنشاء رتبة
    rank = Rank.objects.create(
        name='جندي',
        order=1,
        is_officer=False
    )
    
    # إنشاء حالات خدمية
    status_active = ServiceStatus.objects.create(
        name='عامل',
        classification='active_full',
        requires_document=False
    )
    
    status_absent = ServiceStatus.objects.create(
        name='منقطع',
        classification='inactive_temp',
        requires_document=False
    )
    
    status_martyr = ServiceStatus.objects.create(
        name='شهيد',
        classification='inactive_perm',
        requires_document=True
    )
    
    # إنشاء موقع
    location = GeoDistrict.objects.create(
        name='بغداد',
        code='BGD'
    )
    
    # إنشاء مؤهل
    qualification = Qualification.objects.create(
        name='ابتدائية',
        level=1
    )
    
    # إنشاء أفراد
    personnel_list = []
    for i in range(5):
        personnel = PersonnelMaster.objects.create(
            military_number=f'100000{i}',
            full_name=f'اختبار {i}',
            national_id=f'1234567890{i}',
            current_rank=rank,
            current_status=status_active,
            central_department=department,
            current_geo_location=location,
            qualification=qualification,
            date_of_birth='1990-01-01',
            enlistment_date='2010-01-01'
        )
        personnel_list.append(personnel)
    
    return {
        'user': user,
        'central_department': department,
        'rank': rank,
        'status_active': status_active,
        'status_absent': status_absent,
        'status_martyr': status_martyr,
        'location': location,
        'qualification': qualification,
        'personnel_list': personnel_list,
    }


@pytest.mark.django_db
class TestExcelImportService:
    """اختبارات خدمة الاستيراد"""
    
    def test_import_without_changes(self, setup_test_data):
        """اختبار 1: استيراد ملف بدون تغييرات"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. استيراد نفس الملف بدون تعديل
        excel_file.seek(0)
        file_content = excel_file.read()
        
        import_service = ExcelImportService(
            file_content=file_content,
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        report = import_service.process()
        
        # التحقق
        assert report['success'] is True
        assert report['stats']['total_rows'] > 0
        assert report['stats']['changes_detected'] == 0
        assert report['stats']['errors'] == 0
    
    def test_import_with_simple_changes(self, setup_test_data):
        """اختبار 2: استيراد ملف مع تغييرات بسيطة (أخضر)"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. تعديل الملف (تغيير حالة فرد واحد)
        excel_file.seek(0)
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook['القوة العاملة']
        
        # تغيير الحالة في الصف الثاني (أول فرد)
        sheet.cell(row=2, column=6).value = 'منقطع'  # عمود "متغير الشهر"
        
        # حفظ الملف المعدل
        modified_file = BytesIO()
        workbook.save(modified_file)
        modified_file.seek(0)
        
        # 3. استيراد الملف المعدل
        import_service = ExcelImportService(
            file_content=modified_file.read(),
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        report = import_service.process()
        
        # التحقق
        assert report['success'] is True
        assert report['stats']['changes_detected'] == 1
        assert report['stats']['green_changes'] == 1
        assert report['stats']['yellow_changes'] == 0
        assert report['stats']['red_changes'] == 0
        
        # التحقق من إنشاء StagingRecord
        staging_records = StagingRecord.objects.filter(
            upload_batch_id=import_service.batch_id
        )
        assert staging_records.count() == 1
        assert staging_records.first().severity == 'low'
        assert staging_records.first().requires_document is False
    
    def test_import_with_high_severity_changes(self, setup_test_data):
        """اختبار 3: استيراد ملف مع تغييرات عالية الأهمية (أصفر)"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. تعديل الملف (تغيير إلى حالة حساسة)
        excel_file.seek(0)
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook['القوة العاملة']
        
        # تغيير إلى "شهيد" (يتطلب مستند)
        sheet.cell(row=2, column=6).value = 'شهيد'
        
        modified_file = BytesIO()
        workbook.save(modified_file)
        modified_file.seek(0)
        
        # 3. استيراد
        import_service = ExcelImportService(
            file_content=modified_file.read(),
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        report = import_service.process()
        
        # التحقق
        assert report['stats']['changes_detected'] == 1
        assert report['stats']['yellow_changes'] == 1
        
        staging_record = StagingRecord.objects.get(
            upload_batch_id=import_service.batch_id
        )
        assert staging_record.severity == 'high'
        assert staging_record.requires_document is True
    
    def test_import_with_invalid_export_id(self, setup_test_data):
        """اختبار 4: استيراد بمعرف تصدير غير صحيح"""
        data = setup_test_data
        
        fake_export_id = str(uuid.uuid4())
        
        import_service = ExcelImportService(
            file_content=b'fake content',
            export_id=fake_export_id,
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        with pytest.raises(ImportValidationError) as exc_info:
            import_service.process()
        
        assert 'لم يتم العثور على سجل تصدير' in str(exc_info.value)
    
    def test_import_with_invalid_military_number(self, setup_test_data):
        """اختبار 5: استيراد مع رقم عسكري غير موجود"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. تعديل الملف (تغيير رقم عسكري إلى رقم غير موجود)
        excel_file.seek(0)
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook['القوة العاملة']
        
        # تغيير الرقم العسكري (لكن نحتفظ بـ UUID صحيح)
        sheet.cell(row=2, column=1).value = '9999999'
        
        modified_file = BytesIO()
        workbook.save(modified_file)
        modified_file.seek(0)
        
        # 3. استيراد
        import_service = ExcelImportService(
            file_content=modified_file.read(),
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        report = import_service.process()
        
        # التحقق
        assert report['stats']['errors'] > 0
        assert any('غير موجود في قاعدة البيانات' in str(e) for e in report['errors'])
    
    def test_import_with_name_mismatch(self, setup_test_data):
        """اختبار 6: استيراد مع اختلاف في الاسم"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. تعديل الملف (تغيير الاسم)
        excel_file.seek(0)
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook['القوة العاملة']
        
        sheet.cell(row=2, column=2).value = 'اسم مختلف'
        sheet.cell(row=2, column=6).value = 'منقطع'  # تغيير الحالة أيضاً
        
        modified_file = BytesIO()
        workbook.save(modified_file)
        modified_file.seek(0)
        
        # 3. استيراد
        import_service = ExcelImportService(
            file_content=modified_file.read(),
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        report = import_service.process()
        
        # التحقق
        assert report['stats']['warnings'] > 0
        assert any('اختلاف في الاسم' in str(w) for w in report['warnings'])
        
        staging_record = StagingRecord.objects.get(
            upload_batch_id=import_service.batch_id
        )
        assert staging_record.name_mismatch is True
    
    def test_import_updates_export_log_status(self, setup_test_data):
        """اختبار 7: التحقق من تحديث حالة ExportLog"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # التحقق من الحالة الأولية
        export_log = ExportLog.objects.get(export_id=export_service.export_log.export_id)
        assert export_log.status == 'pending'
        
        # 2. استيراد
        excel_file.seek(0)
        import_service = ExcelImportService(
            file_content=excel_file.read(),
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        import_service.process()
        
        # التحقق من تحديث الحالة
        export_log.refresh_from_db()
        assert export_log.status == 'returned'
    
    def test_import_creates_central_department_compliance(self, setup_test_data):
        """اختبار 8: التحقق من إنشاء سجل التزام الإدارة"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. استيراد
        excel_file.seek(0)
        import_service = ExcelImportService(
            file_content=excel_file.read(),
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        import_service.process()
        
        # التحقق
        compliance = DirectorateCompliance.objects.get(
            central_department=data['central_department'],
            service_month='2026-03'
        )
        assert compliance is not None
        assert compliance.submitted_at is not None
    
    def test_import_multiple_changes(self, setup_test_data):
        """اختبار 9: استيراد مع تغييرات متعددة"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. تعديل الملف (تغيير عدة أفراد)
        excel_file.seek(0)
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook['القوة العاملة']
        
        # تغيير 3 أفراد
        sheet.cell(row=2, column=6).value = 'منقطع'
        sheet.cell(row=3, column=6).value = 'منقطع'
        sheet.cell(row=4, column=6).value = 'شهيد'
        
        modified_file = BytesIO()
        workbook.save(modified_file)
        modified_file.seek(0)
        
        # 3. استيراد
        import_service = ExcelImportService(
            file_content=modified_file.read(),
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        report = import_service.process()
        
        # التحقق
        assert report['stats']['changes_detected'] == 3
        assert report['stats']['green_changes'] == 2  # منقطع × 2
        assert report['stats']['yellow_changes'] == 1  # شهيد
        
        staging_records = StagingRecord.objects.filter(
            upload_batch_id=import_service.batch_id
        )
        assert staging_records.count() == 3
    
    def test_import_helper_function(self, setup_test_data):
        """اختبار 10: اختبار الدالة المساعدة import_service_file"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. استيراد باستخدام الدالة المساعدة
        excel_file.seek(0)
        report = import_service_file(
            file_content=excel_file.read(),
            export_id=str(export_service.export_log.export_id),
            user=data['user'],
            service_month='2026-03'
        )
        
        # التحقق
        assert report['success'] is True
        assert 'batch_id' in report
        assert 'stats' in report
    
    def test_import_with_other_status(self, setup_test_data):
        """اختبار 11: استيراد مع حالة "أخرى" (أحمر)"""
        data = setup_test_data
        
        # 1. تصدير ملف
        export_service = ExcelExportService(
            central_department=data['central_department'],
            service_month='2026-03',
            exported_by=data['user']
        )
        excel_file, filename = export_service.export_and_log()
        
        # 2. تعديل الملف
        excel_file.seek(0)
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook['القوة العاملة']
        
        sheet.cell(row=2, column=6).value = 'أخرى'
        sheet.cell(row=2, column=7).value = 'حالة غير معروفة'
        
        modified_file = BytesIO()
        workbook.save(modified_file)
        modified_file.seek(0)
        
        # 3. استيراد
        import_service = ExcelImportService(
            file_content=modified_file.read(),
            export_id=str(export_service.export_log.export_id),
            imported_by=data['user'],
            service_month='2026-03'
        )
        
        report = import_service.process()
        
        # التحقق
        assert report['stats']['red_changes'] == 1
        
        staging_record = StagingRecord.objects.get(
            upload_batch_id=import_service.batch_id
        )
        assert staging_record.severity == 'high'
        assert staging_record.notes == 'حالة غير معروفة'
